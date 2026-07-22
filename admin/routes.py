"""
admin/routes.py — CarYanams DMS Admin Panel
Full working version with all CRUD for dealers, users, vehicles, leads.
"""

from werkzeug.utils import secure_filename
from flask import (
    Blueprint, current_app, flash, redirect, request, url_for
)
import uuid
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, make_response, Response
from functools import wraps
from datetime import datetime, timedelta, timezone as _tz
import io

_IST = _tz(timedelta(hours=5, minutes=30))


def _now_ist():
    """Return current IST time as a naive datetime."""
    return datetime.now(_IST).replace(tzinfo=None)
import csv
import os
import random
from sqlalchemy import func, case, cast, String

admin_bp = Blueprint('admin', __name__)


# ─── Unique-visitor / unique-actor aggregation helpers ────────────────────────
# These build a single SQL expression that identifies "one person" instead of
# "one row", so COUNT(DISTINCT ...) on it gives real unique users rather than
# inflated page-view / action counts. Nothing about existing row storage
# changes -- every page view and every action is still written exactly as
# before; only how the dashboard stat cards *summarize* those rows changes.
# Built with the `+` string-concat operator (compiles to `||` on both SQLite
# and PostgreSQL) rather than func.concat(), which SQLite has no builtin for.
def _visitor_unique_key_expr():
    """
    Priority order for identifying a unique visitor on VisitorLog rows:
      1. Logged-in user_id (best -- same person across devices counts once)
      2. Fallback: ip_address + device_type + session_id combo
    """
    from models import VisitorLog
    anon_key = (
        func.coalesce(VisitorLog.ip_address, '') + '|' +
        func.coalesce(VisitorLog.device_type, '') + '|' +
        func.coalesce(VisitorLog.session_id, '')
    )
    return case(
        (VisitorLog.user_id.isnot(None), 'u:' + cast(VisitorLog.user_id, String)),
        else_=('a:' + anon_key)
    )


def _activity_unique_key_expr():
    """
    Priority order for identifying a unique actor on AdminLog rows:
      1. user_id (best -- same person across sessions counts once)
      2. Fallback: ip_address + user_role combo
    """
    from models import AdminLog
    anon_key = (
        func.coalesce(AdminLog.ip_address, '') + '|' +
        func.coalesce(AdminLog.user_role, '')
    )
    return case(
        (AdminLog.user_id.isnot(None), 'u:' + cast(AdminLog.user_id, String)),
        else_=('a:' + anon_key)
    )

# ─── Admin Credentials ────────────────────────────────────────────────────────
ADMIN_CREDS = {
    'username': 'admin',
    'password': 'admin123',
    'mobile':   '9876543210',
}

OTP_STORE = {}

GENERAL_SETTINGS = {
    'company_name':  'CarYanams DMS',
    'support_email': 'admin@caryanams.com',
    'contact':       '+91 98765 43210',
    'timezone':      'Asia/Kolkata (IST)',
    'currency':      'INR (₹)',
    'status':        'Active',
    'address':       'Ahmedabad, Gujarat 380001',
}

EMAIL_SETTINGS = {
    'smtp_host':  'smtp.gmail.com',
    'smtp_port':  '587',
    'from_email': 'noreply@caryanams.com',
    'smtp_pass':  '',
    'encryption': 'TLS',
}

# ─── KYC Image Upload Config ──────────────────────────────────────────────────
ALLOWED_EXTENSIONS = {'jpg', 'jpeg', 'png', 'gif', 'webp', 'pdf'}
MAX_KYC_IMAGE_SIZE = 5 * 1024 * 1024  # 5 MB


# ─── Auth Decorators ──────────────────────────────────────────────────────────

def admin_login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_logged_in') and not session.get('sub_admin_logged_in'):
            flash('Please log in to access the admin panel.', 'error')
            return redirect(url_for('admin.login'))
        return f(*args, **kwargs)
    return decorated


def admin_api_login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_logged_in') and not session.get('sub_admin_logged_in'):
            return jsonify({'success': False, 'message': 'Unauthorized'}), 401
        return f(*args, **kwargs)
    return decorated


def is_super_admin():
    return bool(session.get('admin_logged_in'))


def sub_admin_has_perm(perm):
    if is_super_admin():
        return True
    return perm in session.get('sub_admin_permissions', [])


def require_permission(perm):
    """Decorator: page route — 403 if sub-admin lacks permission."""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not session.get('admin_logged_in') and not session.get('sub_admin_logged_in'):
                flash('Please log in to access the admin panel.', 'error')
                return redirect(url_for('admin.login'))
            if not is_super_admin() and perm not in session.get('sub_admin_permissions', []):
                return render_template('admin/403.html', missing_perm=perm,
                                       current_perms=session.get('sub_admin_permissions', [])), 403
            return f(*args, **kwargs)
        return decorated
    return decorator


def require_permission_api(perm):
    """Decorator: API/JSON route — 403 JSON if sub-admin lacks permission."""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not session.get('admin_logged_in') and not session.get('sub_admin_logged_in'):
                return jsonify({'success': False, 'message': 'Unauthorized'}), 401
            if not is_super_admin() and perm not in session.get('sub_admin_permissions', []):
                return jsonify({'success': False, 'message': f'Access Denied: {perm} permission required'}), 403
            return f(*args, **kwargs)
        return decorated
    return decorator


def super_admin_only(f):
    """Decorator: page route — only Super Admin allowed."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_logged_in') and not session.get('sub_admin_logged_in'):
            flash('Please log in to access the admin panel.', 'error')
            return redirect(url_for('admin.login'))
        if not is_super_admin():
            return render_template('admin/403.html', missing_perm='super_admin',
                                   current_perms=session.get('sub_admin_permissions', [])), 403
        return f(*args, **kwargs)
    return decorated


def super_admin_only_api(f):
    """Decorator: API route — only Super Admin allowed."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_logged_in') and not session.get('sub_admin_logged_in'):
            return jsonify({'success': False, 'message': 'Unauthorized'}), 401
        if not is_super_admin():
            return jsonify({'success': False, 'message': 'Access Denied: Super Admin only'}), 403
        return f(*args, **kwargs)
    return decorated


def _resolve_role_and_user():
    """Return (user_role, username) from the current session."""
    if session.get('admin_logged_in'):
        return 'Super Admin', session.get('admin_username', 'admin')
    if session.get('sub_admin_logged_in'):
        return 'Sub Admin', session.get('sub_admin_name', session.get('sub_admin_username', 'subadmin'))
    return 'Admin', session.get('admin_username', 'admin')


def _current_actor_id():
    """Best-effort numeric id of the currently logged-in admin/sub-admin/dealer."""
    if session.get('admin_logged_in'):
        return None  # built-in Super Admin has no row in users/sub_admins
    if session.get('sub_admin_logged_in'):
        return session.get('sub_admin_id')
    uid = session.get('user_id')
    return uid


def log_admin_action(action, module, status='Success', description=None):
    """
    Log an admin/dealer/sub-admin action into the unified AdminLog table.
    Backward compatible: existing calls with just (action, module[, status])
    keep working unchanged. `description` is optional and defaults to `action`.
    """
    try:
        from extensions import db
        from models import AdminLog
        from utils.request_meta import get_request_meta
        user_role, username = _resolve_role_and_user()
        ip, browser, os_name, device = get_request_meta(request)
        log = AdminLog(
            user_id=_current_actor_id(),
            admin_user=username,
            user_role=user_role,
            action=action,
            module=module,
            description=description or action,
            ip_address=ip,
            device=device,
            browser=browser,
            timezone='Asia/Kolkata (IST)',
            status=status,
        )
        db.session.add(log)
        db.session.commit()
    except Exception:
        pass


# ─── Auth Routes ──────────────────────────────────────────────────────────────

@admin_bp.route('/login', methods=['GET', 'POST'])
def login():
    if session.get('admin_logged_in'):
        return redirect(url_for('admin.dashboard'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        if username == ADMIN_CREDS['username'] and password == ADMIN_CREDS['password']:
            session['admin_logged_in'] = True
            session['admin_username'] = username
            log_admin_action('Admin logged in', 'Auth')
            flash('Welcome back, Super Admin!', 'success')
            return redirect(url_for('admin.dashboard'))
        else:
            # Log failed login attempt (best-effort; username is what was typed,
            # not a valid account, so it's recorded as-is for audit purposes)
            try:
                from extensions import db
                from models import AdminLog
                from utils.request_meta import get_request_meta
                ip, browser, os_name, device = get_request_meta(request)
                db.session.add(AdminLog(
                    user_id=None,
                    admin_user=username or 'unknown',
                    user_role='Admin',
                    action='Failed admin login attempt',
                    module='Auth',
                    description=f'Failed login attempt for username "{username}"',
                    ip_address=ip,
                    device=device,
                    browser=browser,
                    timezone='Asia/Kolkata (IST)',
                    status='Failed',
                ))
                db.session.commit()
            except Exception:
                pass
            flash('Invalid username or password.', 'error')
    return render_template('admin/admin_login.html')


@admin_bp.route('/logout')
def logout():
    log_admin_action('Admin logged out', 'Auth')
    for k in ['admin_logged_in','admin_username','sub_admin_logged_in','sub_admin_id',
              'sub_admin_username','sub_admin_name','sub_admin_permissions']:
        session.pop(k, None)
    flash('You have been logged out.', 'success')
    return redirect(url_for('admin.login'))


# ─── Dashboard ────────────────────────────────────────────────────────────────

@admin_bp.route('/')
@admin_login_required
def dashboard():
    from models import User, Vehicle, Lead, Deal, Inquiry, AdminLog
    total_dealers = User.query.filter_by(role='dealer').count()
    total_users = User.query.filter_by(role='user').count()
    total_vehicles = Vehicle.query.count()
    available_v = Vehicle.query.filter_by(status='available').count()
    sold_v = Vehicle.query.filter_by(status='sold').count()
    total_leads = Lead.query.count()
    total_deals = Deal.query.count()
    total_inquiries = Inquiry.query.count()
    paid_deals = Deal.query.filter_by(payment_mode='cash').all()
    monthly_revenue = sum(d.final_price or 0 for d in paid_deals)
    pending_vehicles = Vehicle.query.filter_by(
        approval_status='pending').count()

    logs = AdminLog.query.order_by(AdminLog.created_at.desc()).limit(10).all()

    stats = {
        'total_dealers':    total_dealers,
        'total_users':      total_users,
        'total_vehicles':   total_vehicles,
        'available':        available_v,
        'sold':             sold_v,
        'total_leads':      total_leads,
        'total_deals':      total_deals,
        'total_inquiries':  total_inquiries,
        'monthly_revenue':  monthly_revenue,
        'vehicles_pending': pending_vehicles,
    }
    recent_dealers = User.query.filter_by(role='dealer').order_by(
        User.created_at.desc()).limit(5).all()
    recent_deals = Deal.query.order_by(Deal.created_at.desc()).limit(5).all()

    return render_template('admin/dashboard.html', stats=stats, logs=logs, page='dashboard',
                           recent_dealers=recent_dealers, recent_deals=recent_deals)


# ─── Dealers ──────────────────────────────────────────────────────────────────

@admin_bp.route('/dealers')
@require_permission('dealers')
def all_dealers():
    from models import User
    dealers = User.query.filter_by(role='dealer').order_by(
        User.created_at.desc()).all()
    return render_template('admin/dealers.html', dealers=dealers, page='dealers')


@admin_bp.route('/dealers/<int:dealer_id>')
@require_permission('dealers')
def view_dealer(dealer_id):
    from models import User, Vehicle, Lead, Deal, DealerSubscription, DealerPayment
    dealer = User.query.get_or_404(dealer_id)
    vehicles = Vehicle.query.filter_by(dealer_id=dealer_id).all()
    leads = Lead.query.filter_by(dealer_id=dealer_id).all()
    deals = Deal.query.filter_by(dealer_id=dealer_id).all()
    leads_count = len(leads)
    deals_count = len(deals)
    revenue = sum(d.final_price or 0 for d in deals)

    # ── Subscription & Payment history (demo payment system) ───────────────
    current_dealer_sub = (DealerSubscription.query
                           .filter_by(dealer_id=dealer_id, is_active=True)
                           .order_by(DealerSubscription.activated_at.desc())
                           .first())
    subscription_history = (DealerSubscription.query
                             .filter_by(dealer_id=dealer_id)
                             .order_by(DealerSubscription.activated_at.desc())
                             .all())
    latest_payment = (DealerPayment.query
                       .filter_by(dealer_id=dealer_id)
                       .order_by(DealerPayment.created_at.desc())
                       .first())

    return render_template('admin/view_dealer.html', dealer=dealer, vehicles=vehicles,
                           leads=leads, leads_count=leads_count, deals_count=deals_count,
                           revenue=revenue, page='dealers',
                           current_dealer_sub=current_dealer_sub,
                           subscription_history=subscription_history,
                           latest_payment=latest_payment)


@admin_bp.route('/dealers/add', methods=['GET', 'POST'])
@require_permission('dealers')
def add_dealer():
    from models import User
    from extensions import db
    from werkzeug.security import generate_password_hash
    if request.method == 'POST':
        name = request.form.get('dealer_name', '').strip()
        email = request.form.get('email', '').strip()
        phone = request.form.get('contact', '').strip()
        company = request.form.get('company_name', '').strip()
        city = request.form.get('city', '').strip()
        gst = request.form.get('gst', '').strip()
        plan = request.form.get('plan', 'starter')
        password = request.form.get('password', 'dealer123').strip()
        is_active = request.form.get('is_active', '1') == '1'

        if not name or not email:
            flash('Dealer name and email are required.', 'error')
            return render_template('admin/add_dealer.html', form=request.form, page='dealers')

        if User.query.filter_by(email=email).first():
            flash('A user with this email already exists.', 'error')
            return render_template('admin/add_dealer.html', form=request.form, page='dealers')

        dealer = User(
            name=name, email=email, phone=phone,
            company_name=company, city=city, gst_number=gst,
            role='dealer', subscription_plan=plan, is_active=is_active,
            password_hash=generate_password_hash(password)
        )
        db.session.add(dealer)
        db.session.flush()   # get dealer.id without committing
        # Assign a PERMANENT display_id immediately — it will never change
        from models import generate_display_id
        dealer.display_id = generate_display_id('dealer')
        db.session.commit()
        # reassign_display_ids now only fills gaps for records missing IDs; it
        # will NOT overwrite dealer.display_id we just set.
        try:
            from db import reassign_display_ids
            reassign_display_ids(role='dealer')
        except Exception:
            pass
        log_admin_action(f"Added new dealer {name} [{dealer.display_id}]", 'Dealers')
        flash(f'Dealer "{name}" added successfully! Dealer ID: {dealer.display_id}', 'success')
        return redirect(url_for('admin.all_dealers'))

    return render_template('admin/add_dealer.html', form={}, page='dealers')


@admin_bp.route('/dealers/<int:dealer_id>/edit', methods=['GET', 'POST'])
@require_permission('dealers')
def edit_dealer(dealer_id):
    from models import User
    from extensions import db
    dealer = User.query.get_or_404(dealer_id)
    if request.method == 'POST':
        dealer.name = request.form.get('dealer_name', dealer.name).strip()
        dealer.phone = request.form.get('contact', dealer.phone or '').strip()
        dealer.email = request.form.get('email', dealer.email).strip()
        dealer.company_name = request.form.get(
            'company_name', dealer.company_name or '').strip()
        dealer.city = request.form.get('city', dealer.city or '').strip()
        dealer.gst_number = request.form.get(
            'gst', dealer.gst_number or '').strip()
        dealer.address = request.form.get(
            'address', dealer.address or '').strip()
        plan = request.form.get('plan', '').strip()
        if plan:
            dealer.subscription_plan = plan
        status_val = request.form.get('is_active', '')
        if status_val != '':
            dealer.is_active = (status_val == '1')
        db.session.commit()
        log_admin_action(f"Updated dealer {dealer.name}", 'Dealers')
        flash('Dealer updated successfully!', 'success')
        return redirect(url_for('admin.view_dealer', dealer_id=dealer_id))

    return render_template('admin/edit_dealer.html', dealer=dealer, form={}, page='dealers')


@admin_bp.route('/api/dealers/<int:dealer_id>/toggle-status', methods=['POST'])
@require_permission_api('dealers')
def toggle_dealer_status(dealer_id):
    from models import User
    from extensions import db
    dealer = User.query.get_or_404(dealer_id)
    dealer.is_active = not dealer.is_active
    db.session.commit()
    status = 'Active' if dealer.is_active else 'Suspended'
    log_admin_action(
        f"{'Activated' if dealer.is_active else 'Suspended'} dealer {dealer.name}", 'Dealers')
    return jsonify({'success': True, 'status': status, 'is_active': dealer.is_active})


@admin_bp.route('/api/dealers/<int:dealer_id>/set-status', methods=['POST'])
@require_permission_api('dealers')
def set_dealer_status(dealer_id):
    from models import User
    from extensions import db
    dealer = User.query.get_or_404(dealer_id)
    data = request.get_json(silent=True) or {}
    new_status = data.get('status', 'active')
    dealer.is_active = (new_status == 'active')
    db.session.commit()
    log_admin_action(f"Set dealer {dealer.name} to {new_status}", 'Dealers')
    return jsonify({'success': True, 'status': 'Active' if dealer.is_active else 'Suspended'})


@admin_bp.route('/api/dealers/<int:dealer_id>/delete', methods=['POST'])
@require_permission_api('dealers')
def delete_dealer(dealer_id):
    from models import User, Vehicle, VehicleImage
    from extensions import db
    import shutil

    dealer = User.query.get_or_404(dealer_id)
    name = dealer.name

    # ── Step 1: Delete all vehicle images from DISK ───────────────────────────
    # (DB records will cascade-delete when dealer is deleted via model relationship)
    try:
        dealer_vehicles = Vehicle.query.filter_by(dealer_id=dealer_id).all()
        for vehicle in dealer_vehicles:
            # Delete individual VehicleImage files
            vehicle_imgs = VehicleImage.query.filter_by(vehicle_id=vehicle.id).all()
            for vi in vehicle_imgs:
                for base in [
                    os.path.join(current_app.root_path, 'static', 'images', 'uploads'),
                    os.path.join(current_app.root_path, 'static', 'uploads',
                                 'vehicles', str(vehicle.id)),
                ]:
                    img_path = os.path.join(base, vi.filename)
                    if os.path.exists(img_path):
                        try:
                            os.remove(img_path)
                        except OSError:
                            pass

            # Delete primary image file
            if vehicle.image_filename:
                img_path = os.path.join(
                    current_app.root_path, 'static', 'images', 'uploads',
                    vehicle.image_filename)
                if os.path.exists(img_path):
                    try:
                        os.remove(img_path)
                    except OSError:
                        pass

            # Delete vehicle upload folder
            veh_folder = os.path.join(
                current_app.root_path, 'static', 'uploads', 'vehicles', str(vehicle.id))
            if os.path.isdir(veh_folder):
                try:
                    shutil.rmtree(veh_folder)
                except OSError:
                    pass

        # Delete dealer KYC documents folder
        dealer_folder = os.path.join(
            current_app.root_path, 'static', 'uploads', 'dealers', str(dealer_id))
        if os.path.isdir(dealer_folder):
            try:
                shutil.rmtree(dealer_folder)
            except OSError:
                pass

    except Exception as e:
        current_app.logger.warning(
            f"delete_dealer: image cleanup error for dealer {dealer_id}: {e}")

    # ── Step 2: Delete dealer from DB (cascades to vehicles, leads, deals etc.) ─
    deleted_display_id = dealer.display_id  # capture before delete for log
    db.session.delete(dealer)
    db.session.commit()

    # PERMANENT ID SYSTEM: We do NOT call reassign_display_ids after a delete.
    # The deleted dealer's ID (e.g. D3) is retired forever and will never be
    # reused for any new dealer. The next new dealer will get max+1.
    # reassign_display_ids is now assign-only and does not renumber existing records.

    log_admin_action(f"Deleted dealer {name} [{deleted_display_id}] (id={dealer_id}) and all related data", 'Dealers')
    return jsonify({'success': True, 'message': f'Dealer "{name}" [{deleted_display_id}] and all associated data have been deleted successfully.'})


@admin_bp.route('/api/dealers/<int:dealer_id>/update-subscription', methods=['POST'])
@require_permission_api('dealers')
def update_dealer_subscription(dealer_id):
    from models import User
    from extensions import db
    from datetime import timedelta
    dealer = User.query.get_or_404(dealer_id)
    data = request.get_json(silent=True) or request.form
    plan = data.get('plan', dealer.subscription_plan)
    days = int(data.get('days', 365))
    dealer.subscription_plan = plan
    dealer.subscription_expiry = _now_ist() + timedelta(days=days)
    dealer.subscription_status = 'active'
    db.session.commit()
    log_admin_action(
        f"Updated subscription for {dealer.name} to {plan}", 'Dealers')
    return jsonify({'success': True, 'message': f'Subscription updated to {plan} for {days} days.'})


@admin_bp.route('/dealers/<int:dealer_id>/update-subscription', methods=['POST'])
@require_permission('dealers')
def update_dealer_subscription_form(dealer_id):
    from models import User
    from extensions import db
    from datetime import timedelta
    dealer = User.query.get_or_404(dealer_id)
    plan = request.form.get('plan', dealer.subscription_plan)
    days = int(request.form.get('days', 365))
    dealer.subscription_plan = plan
    dealer.subscription_expiry = _now_ist() + timedelta(days=days)
    dealer.subscription_status = 'active'
    db.session.commit()
    log_admin_action(
        f"Updated subscription for {dealer.name} to {plan}", 'Dealers')
    flash(f'Subscription updated to {plan} for {days} days.', 'success')
    return redirect(url_for('admin.view_dealer', dealer_id=dealer_id))


# ─── KYC / Dealer Requests ───────────────────────────────────────────────────

@admin_bp.route('/dealer-requests')
@require_permission('dealers')
def dealer_requests():
    from models import User
    pending = User.query.filter_by(role='dealer', is_active=False).all()
    approved = User.query.filter_by(role='dealer', is_active=True).all()
    return render_template('admin/dealer_requests.html', pending=pending,
                           approved=approved, page='dealer_requests')

@admin_bp.route('/api/dealers/<int:dealer_id>/approve-kyc', methods=['POST'])
@require_permission_api('dealers')
def approve_kyc(dealer_id):
    from models import User, DealerKYC, DealerNotification
    from extensions import db
    from datetime import datetime
    dealer = User.query.get_or_404(dealer_id)
    kyc = DealerKYC.query.filter_by(dealer_id=dealer_id).first()
    reviewer = session.get('admin_username') or session.get('sub_admin_username') or 'Admin'
    now = _now_ist()
    if kyc:
        for doc in ('aadhaar_front', 'aadhaar_back', 'pan_card'):
            setattr(kyc, doc + '_status', 'approved')
            setattr(kyc, doc + '_reject', None)
            setattr(kyc, doc + '_reviewed_by', reviewer)
            setattr(kyc, doc + '_reviewed_at', now)
        kyc.kyc_status = 'approved'
        kyc.reviewed_at = now
        kyc.reviewed_by = reviewer
        kyc.rejection_reason = None
        dealer.is_active = True
    db.session.add(DealerNotification(
        dealer_id=dealer_id,
        title='KYC Approved — Account Activated',
        message='Your KYC verification has been approved. Your dealer account is now fully active and all features are unlocked.',
        notif_type='success'
    ))
    db.session.commit()
    log_admin_action(f"Approved KYC for dealer {dealer.name}", 'Dealers')
    return jsonify({'success': True, 'message': f'KYC approved for {dealer.name}'})


@admin_bp.route('/api/dealers/<int:dealer_id>/reject-kyc', methods=['POST'])
@require_permission_api('dealers')
def reject_kyc(dealer_id):
    from models import User, DealerKYC, DealerNotification
    from extensions import db
    from datetime import datetime
    dealer = User.query.get_or_404(dealer_id)
    data = request.get_json(silent=True) or {}
    reason = data.get('reason', '')
    reviewer = session.get('admin_username') or session.get('sub_admin_username') or 'Admin'
    now = _now_ist()
    dealer.is_active = False
    kyc = DealerKYC.query.filter_by(dealer_id=dealer_id).first()
    if kyc:
        kyc.kyc_status = 'rejected'
        kyc.reviewed_at = now
        kyc.reviewed_by = reviewer
        kyc.rejection_reason = reason
        for doc in ('aadhaar_front', 'aadhaar_back', 'pan_card'):
            if (getattr(kyc, doc + '_status') or 'pending') == 'pending':
                setattr(kyc, doc + '_status', 'rejected')
                setattr(kyc, doc + '_reject', reason)
                setattr(kyc, doc + '_reviewed_by', reviewer)
                setattr(kyc, doc + '_reviewed_at', now)
    db.session.add(DealerNotification(
        dealer_id=dealer_id,
        title='KYC Application Rejected',
        message=f'Your KYC application has been rejected. Reason: {reason}' if reason else 'Your KYC application has been rejected. Please re-upload your documents.',
        notif_type='danger'
    ))
    db.session.commit()
    log_admin_action(
        f"Rejected KYC for dealer {dealer.name}: {reason}", 'Dealers')
    return jsonify({'success': True, 'message': f'KYC rejected for {dealer.name}'})


@admin_bp.route('/api/kyc/<int:dealer_id>/doc-approve', methods=['POST'])
@require_permission_api('kyc')
def kyc_doc_approve(dealer_id):
    """Approve a single KYC document."""
    from models import User, DealerKYC, DealerNotification
    from extensions import db
    from datetime import datetime
    data = request.get_json(silent=True) or {}
    doc_key = data.get('doc_key')
    if doc_key not in ('aadhaar_front', 'aadhaar_back', 'pan_card'):
        return jsonify({'success': False, 'message': 'Invalid document key'})
    kyc = DealerKYC.query.filter_by(dealer_id=dealer_id).first()
    if not kyc:
        return jsonify({'success': False, 'message': 'KYC record not found'})
    reviewer = session.get('admin_username') or session.get('sub_admin_username') or 'Admin'
    now = _now_ist()
    prev_doc_status = getattr(kyc, doc_key + '_status') or 'pending'
    setattr(kyc, doc_key + '_status', 'approved')
    setattr(kyc, doc_key + '_reject', None)
    setattr(kyc, doc_key + '_reviewed_by', reviewer)
    setattr(kyc, doc_key + '_reviewed_at', now)
    kyc.reviewed_at = now
    kyc.reviewed_by = reviewer
    kyc.recalculate_status()
    _log_kyc_review(dealer_id, doc_key, 'approved', None, prev_doc_status, reviewer)
    doc_labels = {'aadhaar_front': 'Aadhaar Front', 'aadhaar_back': 'Aadhaar Back', 'pan_card': 'PAN Card'}
    if kyc.kyc_status == 'approved':
        dealer = User.query.get(dealer_id)
        if dealer:
            dealer.is_active = True
        db.session.add(DealerNotification(
            dealer_id=dealer_id,
            title='KYC Approved — Account Activated',
            message='All your KYC documents have been verified. Your dealer account is now fully active.',
            notif_type='success'
        ))
    else:
        db.session.add(DealerNotification(
            dealer_id=dealer_id,
            title=f'{doc_labels[doc_key]} Approved',
            message=f'Your {doc_labels[doc_key]} document has been approved by the admin.',
            notif_type='success'
        ))
    db.session.commit()
    return jsonify({'success': True, 'message': f'{doc_labels[doc_key]} approved',
                    'overall_status': kyc.kyc_status, 'counts': _kyc_counts()})


@admin_bp.route('/api/kyc/<int:dealer_id>/doc-reject', methods=['POST'])
@require_permission_api('kyc')
def kyc_doc_reject(dealer_id):
    """Reject a single KYC document."""
    from models import User, DealerKYC, DealerNotification
    from extensions import db
    from datetime import datetime
    data = request.get_json(silent=True) or {}
    doc_key = data.get('doc_key')
    reason = data.get('reason', '')
    if doc_key not in ('aadhaar_front', 'aadhaar_back', 'pan_card'):
        return jsonify({'success': False, 'message': 'Invalid document key'})
    if not reason:
        return jsonify({'success': False, 'message': 'Rejection reason required'})
    kyc = DealerKYC.query.filter_by(dealer_id=dealer_id).first()
    if not kyc:
        return jsonify({'success': False, 'message': 'KYC record not found'})
    reviewer = session.get('admin_username') or session.get('sub_admin_username') or 'Admin'
    now = _now_ist()
    prev_doc_status = getattr(kyc, doc_key + '_status') or 'pending'
    setattr(kyc, doc_key + '_status', 'rejected')
    setattr(kyc, doc_key + '_reject', reason)
    setattr(kyc, doc_key + '_reviewed_by', reviewer)
    setattr(kyc, doc_key + '_reviewed_at', now)
    kyc.reviewed_at = now
    kyc.reviewed_by = reviewer
    kyc.recalculate_status()
    _log_kyc_review(dealer_id, doc_key, 'rejected', reason, prev_doc_status, reviewer)
    doc_labels = {'aadhaar_front': 'Aadhaar Front', 'aadhaar_back': 'Aadhaar Back', 'pan_card': 'PAN Card'}
    db.session.add(DealerNotification(
        dealer_id=dealer_id,
        title=f'{doc_labels[doc_key]} Rejected',
        message=f'Your {doc_labels[doc_key]} document was rejected. Reason: {reason}. Please re-upload this document.',
        notif_type='danger'
    ))
    db.session.commit()
    return jsonify({'success': True, 'message': f'{doc_labels[doc_key]} rejected',
                    'overall_status': kyc.kyc_status, 'counts': _kyc_counts()})


# ─── KYC Review Audit CRUD ────────────────────────────────────────────────────

@admin_bp.route('/api/kyc/<int:dealer_id>/reviews', methods=['GET'])
@require_permission_api('kyc')
def api_kyc_reviews(dealer_id):
    """Return active (non-deleted) review history for a dealer."""
    from models import KYCReview
    reviews = (KYCReview.query
               .filter_by(dealer_id=dealer_id)
               .filter(KYCReview.deleted_at.is_(None))
               .order_by(KYCReview.reviewed_at.desc())
               .all())
    return jsonify({'success': True, 'reviews': [r.to_dict() for r in reviews]})


@admin_bp.route('/api/kyc/reviews/<int:review_id>', methods=['DELETE'])
@require_permission_api('kyc')
def api_delete_kyc_review(review_id):
    """Soft-delete a single KYC review record (keeps audit trail)."""
    from models import KYCReview
    from extensions import db
    review = KYCReview.query.get_or_404(review_id)
    review.soft_delete()
    db.session.commit()
    log_admin_action(f'Soft-deleted KYC review id={review_id}', 'KYC')
    return jsonify({'success': True, 'message': 'Review record removed.'})


# ─── Users ────────────────────────────────────────────────────────────────────

@admin_bp.route('/users')
@require_permission('users')
def all_users():
    from models import User
    users = User.query.filter_by(role='user').order_by(
        User.created_at.desc()).all()

    # Export support
    export_fmt = request.args.get('export', '').strip()
    if export_fmt in ('csv', 'xlsx'):
        headers_row = ['ID', 'Name', 'Email', 'Phone', 'City', 'Status', 'Joined']
        rows_data = []
        for u in users:
            rows_data.append([
                u.display_id or ('U' + str(u.id)),
                u.name,
                u.email,
                u.phone or '',
                u.city or '',
                'Active' if u.is_active else 'Blocked',
                u.created_at.strftime('%Y-%m-%d') if u.created_at else ''
            ])
        if export_fmt == 'xlsx':
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = 'Users'
                header_fill = PatternFill('solid', fgColor='0369A1')
                header_font = Font(bold=True, color='FFFFFF')
                ws.append(headers_row)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                for row in rows_data:
                    ws.append(row)
                for col in ws.columns:
                    max_len = max((len(str(c.value or '')) for c in col), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                output = make_response(buf.read())
                output.headers['Content-Disposition'] = 'attachment; filename=users_export.xlsx'
                output.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                return output
            except ImportError:
                flash('openpyxl is required for Excel export.', 'error')
                return redirect(url_for('admin.all_users'))
        else:
            si = io.StringIO()
            cw = csv.writer(si)
            cw.writerow(headers_row)
            for row in rows_data:
                cw.writerow(row)
            output = make_response(si.getvalue())
            output.headers['Content-Disposition'] = 'attachment; filename=users_export.csv'
            output.headers['Content-type'] = 'text/csv'
            return output

    return render_template('admin/users.html', users=users, page='users')


@admin_bp.route('/users/add', methods=['GET', 'POST'])
@require_permission('users')
def add_user():
    from models import User
    from extensions import db
    from werkzeug.security import generate_password_hash
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        city = request.form.get('city', '').strip()
        password = request.form.get('password', 'user123').strip()
        is_active = request.form.get('is_active', '1') == '1'

        if not name or not email:
            flash('Name and email are required.', 'error')
            return render_template('admin/add_user.html', form=request.form, page='users')

        if User.query.filter_by(email=email).first():
            flash('A user with this email already exists.', 'error')
            return render_template('admin/add_user.html', form=request.form, page='users')

        user = User(
            name=name, email=email, phone=phone, city=city,
            role='user', is_active=is_active,
            password_hash=generate_password_hash(password)
        )
        db.session.add(user)
        db.session.commit()
        # Reassign all user IDs to keep series gapless (U1, U2, U3 …)
        try:
            from db import reassign_display_ids
            reassign_display_ids(role='user')
        except Exception:
            pass
        log_admin_action(f"Added new user {name}", 'Users')
        flash(f'User "{name}" added successfully!', 'success')
        return redirect(url_for('admin.all_users'))

    return render_template('admin/add_user.html', form={}, page='users')


@admin_bp.route('/users/<int:user_id>')
@require_permission('users')
def view_user(user_id):
    from models import User, Lead, Inquiry
    user = User.query.get_or_404(user_id)
    inquiries = Inquiry.query.filter_by(dealer_id=user_id).all()
    leads = Lead.query.filter_by(dealer_id=user_id).all()
    return render_template('admin/view_user.html', user=user,
                           inquiries=inquiries, leads=leads, page='users')


@admin_bp.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@require_permission('users')
def edit_user(user_id):
    from models import User
    from extensions import db
    user = User.query.get_or_404(user_id)
    if request.method == 'POST':
        user.name = request.form.get('name', user.name).strip()
        user.email = request.form.get('email', user.email).strip()
        user.phone = request.form.get('phone', user.phone or '').strip()
        user.city = request.form.get('city', user.city or '').strip()
        status_val = request.form.get('is_active', '')
        if status_val != '':
            user.is_active = (status_val == '1')
        db.session.commit()
        log_admin_action(f"Updated user {user.name}", 'Users')
        flash('User updated successfully!', 'success')
        return redirect(url_for('admin.all_users'))
    return render_template('admin/edit_user.html', user=user, page='users')


@admin_bp.route('/api/users/<int:user_id>/toggle-status', methods=['POST'])
@require_permission_api('users')
def toggle_user_status(user_id):
    from models import User
    from extensions import db
    user = User.query.get_or_404(user_id)
    user.is_active = not user.is_active
    db.session.commit()
    status = 'Active' if user.is_active else 'Blocked'
    log_admin_action(
        f"{'Unblocked' if user.is_active else 'Blocked'} user {user.name}", 'Users')
    return jsonify({'success': True, 'status': status, 'is_active': user.is_active})


@admin_bp.route('/api/users/<int:user_id>/delete', methods=['POST'])
@require_permission_api('users')
def delete_user(user_id):
    from models import User
    from extensions import db
    user = User.query.get_or_404(user_id)
    name = user.name
    db.session.delete(user)
    db.session.commit()

    # Reassign user display IDs so series stays gapless (U1, U2, U3 …)
    try:
        from db import reassign_display_ids
        reassign_display_ids(role='user')
    except Exception:
        pass

    log_admin_action(f"Deleted user {name}", 'Users')
    return jsonify({'success': True, 'message': f'User "{name}" deleted.'})


@admin_bp.route('/api/users/<int:user_id>/convert-to-lead', methods=['POST'])
@require_permission_api('users')
def convert_user_to_lead(user_id):
    """Convert a registered user into a Lead record (admin action)."""
    from models import User, Lead
    from extensions import db
    user = User.query.get_or_404(user_id)

    # Check if a lead with this phone already exists to avoid duplicates
    existing = Lead.query.filter_by(customer_phone=user.phone or '').first() if user.phone else None
    if existing:
        return jsonify({'success': False, 'message': f'A lead for this user already exists (Lead #{existing.id}).'})

    lead = Lead(
        customer_name=user.name,
        customer_email=user.email,
        customer_phone=user.phone or '',
        customer_city=user.city or '',
        source='converted_user',
        stage='new',
        notes=f'Converted from registered user {user.display_id or user.email} on admin panel.',
    )
    db.session.add(lead)
    db.session.commit()
    log_admin_action(f"Converted user {user.name} to lead #{lead.id}", 'Users')
    return jsonify({'success': True, 'message': f'User "{user.name}" converted to lead successfully!', 'lead_id': lead.id})


# ─── Vehicles ─────────────────────────────────────────────────────────────────

@admin_bp.route('/vehicles')
@require_permission('vehicles')
def all_vehicles():
    from models import Vehicle, User
    vehicles = Vehicle.query.order_by(Vehicle.created_at.desc()).all()
    dealers = User.query.filter_by(role='dealer').all()
    return render_template('admin/vehicles.html', vehicles=vehicles,
                           dealers=dealers, page='vehicles')


@admin_bp.route('/vehicles/<int:vehicle_id>')
@require_permission('vehicles')
def view_vehicle(vehicle_id):
    from models import Vehicle, Lead
    vehicle = Vehicle.query.get_or_404(vehicle_id)
    leads = Lead.query.filter_by(vehicle_id=vehicle_id).all()
    return render_template('admin/view_vehicle.html', vehicle=vehicle, leads=leads, page='vehicles')


@admin_bp.route('/vehicles/add', methods=['GET', 'POST'])
@require_permission('vehicles')
def add_vehicle():
    from models import User, Vehicle, VehicleImage
    from extensions import db
    dealers = User.query.filter_by(role='dealer').all()

    if request.method == 'POST':
        dealer_id = request.form.get('dealer_id', type=int)
        make = request.form.get('brand', '').strip()
        model = request.form.get('model', '').strip()
        year = request.form.get('year', type=int)
        fuel_type = request.form.get('fuel_type', '').strip()
        price = request.form.get('price', 0, type=float)
        mileage = request.form.get('mileage', 0, type=int)
        color = request.form.get('color', '').strip()
        reg_number = request.form.get('reg_number', '').strip()
        desc = request.form.get('description', '').strip()
        approval = request.form.get('approval_status', 'approved')
        # new condition detail fields
        accident_history   = request.form.get('accident_history', 'NA').strip()
        loan_status        = request.form.get('loan_status', 'NA').strip()
        rc_service_records = request.form.get('rc_service_records', 'NA').strip()
        major_issues_list  = request.form.getlist('major_issues')
        major_issues       = ','.join(major_issues_list) if major_issues_list else 'None'
        keys_available     = request.form.get('keys_available', 'NA').strip()
        body_panel_status  = request.form.get('body_panel_status', 'NA').strip()

        if not make or not model or not dealer_id:
            flash('Brand, model and dealer are required.', 'error')
            return render_template('admin/add_vehicle.html', dealers=dealers, form=request.form, page='vehicles')

        v = Vehicle(
            dealer_id=dealer_id, make=make, model=model, year=year,
            fuel_type=fuel_type, price=price, mileage=mileage,
            color=color, registration_number=reg_number,
            description=desc, status='available', approval_status=approval,
            accident_history=accident_history,
            loan_status=loan_status,
            rc_service_records=rc_service_records,
            major_issues=major_issues,
            keys_available=keys_available,
            body_panel_status=body_panel_status,
        )
        db.session.add(v)
        db.session.commit()

        try:
            from utils.upload_helpers import save_image, validate_image
            uploaded_images = request.files.getlist('images')
            img_folder = _vehicle_upload_folder()
            os.makedirs(img_folder, exist_ok=True)
            sort_idx = 0
            for f in uploaded_images[:20]:
                if not f or not f.filename:
                    continue
                ok, _ = validate_image(f)
                if not ok:
                    continue
                fname = save_image(f, img_folder, prefix='vimg')
                if fname:
                    vi = VehicleImage(
                        vehicle_id=v.id, filename=fname, sort_order=sort_idx)
                    db.session.add(vi)
                    if sort_idx == 0:
                        v.image_filename = fname
                    sort_idx += 1
            if sort_idx:
                db.session.commit()
        except Exception as e:
            print(f'[add_vehicle] Image upload error: {e}')

        log_admin_action(f"Added vehicle {make} {model}", 'Vehicles')
        flash(f'Vehicle "{make} {model}" added!', 'success')
        return redirect(url_for('admin.view_vehicle', vehicle_id=v.id))

    return render_template('admin/add_vehicle.html', dealers=dealers, form={}, page='vehicles')


@admin_bp.route('/vehicles/<int:vehicle_id>/edit', methods=['GET', 'POST'])
@require_permission('vehicles')
def edit_vehicle(vehicle_id):
    from models import Vehicle, User
    from extensions import db
    vehicle = Vehicle.query.get_or_404(vehicle_id)
    dealers = User.query.filter_by(role='dealer').all()
    if request.method == 'POST':
        vehicle.make = request.form.get('brand', vehicle.make).strip()
        vehicle.model = request.form.get('model', vehicle.model).strip()
        vehicle.year = request.form.get('year', vehicle.year, type=int)
        vehicle.fuel_type = request.form.get(
            'fuel_type', vehicle.fuel_type or '').strip()
        vehicle.price = request.form.get('price', vehicle.price, type=float)
        vehicle.mileage = request.form.get(
            'mileage', vehicle.mileage or 0, type=int)
        vehicle.color = request.form.get('color', vehicle.color or '').strip()
        vehicle.description = request.form.get(
            'description', vehicle.description or '').strip()
        reg_number = request.form.get('reg_number', '').strip()
        if reg_number:
            vehicle.registration_number = reg_number.upper()
        # new condition detail fields
        vehicle.accident_history   = request.form.get('accident_history', vehicle.accident_history or 'NA').strip()
        vehicle.loan_status        = request.form.get('loan_status', vehicle.loan_status or 'NA').strip()
        vehicle.rc_service_records = request.form.get('rc_service_records', vehicle.rc_service_records or 'NA').strip()
        major_issues_list = request.form.getlist('major_issues')
        if major_issues_list:
            vehicle.major_issues = ','.join(major_issues_list)
        vehicle.keys_available    = request.form.get('keys_available', vehicle.keys_available or 'NA').strip()
        vehicle.body_panel_status = request.form.get('body_panel_status', vehicle.body_panel_status or 'NA').strip()
        db.session.commit()
        log_admin_action(
            f"Updated vehicle {vehicle.make} {vehicle.model}", 'Vehicles')
        flash('Vehicle updated successfully!', 'success')
        return redirect(url_for('admin.all_vehicles'))
    return render_template('admin/edit_vehicle.html', vehicle=vehicle, dealers=dealers, page='vehicles')


@admin_bp.route('/api/vehicles/<int:vehicle_id>/delete', methods=['POST'])
@require_permission_api('vehicles')
def delete_vehicle(vehicle_id):
    from models import Vehicle
    from extensions import db
    v = Vehicle.query.get_or_404(vehicle_id)
    label = f"{v.make} {v.model}"
    db.session.delete(v)
    db.session.commit()
    log_admin_action(f"Deleted vehicle {label}", 'Vehicles')
    return jsonify({'success': True, 'message': f'Vehicle "{label}" deleted.'})


@admin_bp.route('/api/vehicles/<int:vehicle_id>/feature', methods=['POST'])
@require_permission_api('vehicles')
def feature_vehicle(vehicle_id):
    from models import Vehicle
    from extensions import db
    v = Vehicle.query.get_or_404(vehicle_id)
    v.featured = not v.featured
    db.session.commit()
    return jsonify({'success': True, 'featured': v.featured})


# ─── Leads ────────────────────────────────────────────────────────────────────

@admin_bp.route('/leads')
@require_permission('leads')
def all_leads():
    from models import Lead, User, Vehicle, Agent
    dealers = User.query.filter_by(role='dealer').all()
    vehicles = Vehicle.query.filter_by(approval_status='approved').all()
    agents = Agent.query.all()

    # Filters
    search   = request.args.get('search', '').strip()
    stage_f  = request.args.get('stage', '').strip()
    dealer_f = request.args.get('dealer_id', '', type=str).strip()
    page     = request.args.get('page', 1, type=int)
    per_page = 20

    # Export (CSV or Excel)
    export_fmt = request.args.get('export', '')
    if export_fmt in ('csv', 'xlsx'):
        query = Lead.query
        if search:
            query = query.filter(
                Lead.customer_name.ilike(f'%{search}%') |
                Lead.customer_phone.ilike(f'%{search}%') |
                Lead.customer_email.ilike(f'%{search}%'))
        if stage_f:
            query = query.filter_by(stage=stage_f)
        if dealer_f:
            query = query.filter_by(dealer_id=int(dealer_f))
        all_leads_export = query.order_by(Lead.created_at.desc()).all()

        headers_row = ['ID','Customer Name','Phone','Email','City','Stage','Source','Budget',
                       'Dealer ID','Dealer','Agent','Notes','Follow-up','Created']
        rows_data = []
        for l in all_leads_export:
            rows_data.append([
                l.id, l.customer_name, l.customer_phone, l.customer_email or '',
                l.customer_city or '', l.stage, l.source or '',
                l.budget or '',
                (l.dealer.display_id if l.dealer and l.dealer.display_id else ('D' + str(l.dealer_id)) if l.dealer_id else ''),
                l.dealer.name if l.dealer else l.dealer_id,
                l.agent.name if l.agent else '',
                (l.notes or '').replace('\n', ' '),
                l.follow_up_date.strftime('%Y-%m-%d %H:%M') if l.follow_up_date else '',
                l.created_at.strftime('%Y-%m-%d') if l.created_at else ''
            ])

        if export_fmt == 'xlsx':
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = 'Leads'
                # Header row styling
                header_fill = PatternFill('solid', fgColor='4F46E5')
                header_font = Font(bold=True, color='FFFFFF')
                ws.append(headers_row)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                for row in rows_data:
                    ws.append(row)
                # Auto-fit columns
                for col in ws.columns:
                    max_len = max((len(str(c.value or '')) for c in col), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                output = make_response(buf.read())
                output.headers['Content-Disposition'] = 'attachment; filename=leads_export.xlsx'
                output.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                return output
            except ImportError:
                flash('openpyxl is required for Excel export. Run: pip install openpyxl', 'error')
                return redirect(url_for('admin.all_leads'))
        else:
            si = io.StringIO()
            cw = csv.writer(si)
            cw.writerow(headers_row)
            for row in rows_data:
                cw.writerow(row)
            output = make_response(si.getvalue())
            output.headers['Content-Disposition'] = 'attachment; filename=leads_export.csv'
            output.headers['Content-type'] = 'text/csv'
            return output

    query = Lead.query
    if search:
        query = query.filter(
            Lead.customer_name.ilike(f'%{search}%') |
            Lead.customer_phone.ilike(f'%{search}%') |
            Lead.customer_email.ilike(f'%{search}%'))
    if stage_f:
        query = query.filter_by(stage=stage_f)
    if dealer_f:
        query = query.filter_by(dealer_id=int(dealer_f))

    total  = query.count()
    pages  = max(1, (total + per_page - 1) // per_page)
    page   = max(1, min(page, pages))
    leads  = query.order_by(Lead.created_at.desc()).offset((page-1)*per_page).limit(per_page).all()

    return render_template('admin/leads.html', leads=leads, dealers=dealers,
                           vehicles=vehicles, agents=agents, page='leads',
                           current_page=page, total_pages=pages, total=total,
                           search=search, stage_filter=stage_f, dealer_filter=dealer_f)


@admin_bp.route('/leads/add', methods=['GET', 'POST'])
@require_permission('leads')
def add_lead():
    from models import Lead, User, Vehicle, Agent
    from extensions import db
    dealers = User.query.filter_by(role='dealer').all()
    vehicles = Vehicle.query.all()
    agents = Agent.query.all()

    if request.method == 'POST':
        dealer_id = request.form.get('dealer_id', type=int)
        customer_name = request.form.get('customer_name', '').strip()
        customer_phone = request.form.get('customer_phone', '').strip()
        customer_email = request.form.get('customer_email', '').strip()
        customer_city = request.form.get('customer_city', '').strip()
        vehicle_id = request.form.get('vehicle_id', type=int) or None
        agent_id = request.form.get('agent_id', type=int) or None
        source = request.form.get('source', 'admin')
        stage = request.form.get('stage', 'new')
        budget = request.form.get('budget', type=float) or None
        notes = request.form.get('notes', '').strip()

        if not customer_name or not customer_phone or not dealer_id:
            flash('Customer name, phone and dealer are required.', 'error')
            return render_template('admin/add_lead.html', dealers=dealers,
                                   vehicles=vehicles, agents=agents, form=request.form, page='leads')

        lead = Lead(
            dealer_id=dealer_id, vehicle_id=vehicle_id, agent_id=agent_id,
            customer_name=customer_name, customer_phone=customer_phone,
            customer_email=customer_email, customer_city=customer_city,
            source=source, stage=stage, budget=budget, notes=notes
        )
        db.session.add(lead)
        db.session.commit()
        log_admin_action(f"Added lead for {customer_name}", 'Leads')
        flash(f'Lead for "{customer_name}" added!', 'success')
        return redirect(url_for('admin.all_leads'))

    return render_template('admin/add_lead.html', dealers=dealers,
                           vehicles=vehicles, agents=agents, form={}, page='leads')


@admin_bp.route('/leads/<int:lead_id>')
@require_permission('leads')
def view_lead(lead_id):
    from models import Lead
    lead = Lead.query.get_or_404(lead_id)
    return render_template('admin/view_lead.html', lead=lead, page='leads')


@admin_bp.route('/leads/<int:lead_id>/edit', methods=['GET', 'POST'])
@require_permission('leads')
def edit_lead(lead_id):
    from models import Lead, User, Vehicle, Agent
    from extensions import db
    lead = Lead.query.get_or_404(lead_id)
    dealers = User.query.filter_by(role='dealer').all()
    vehicles = Vehicle.query.all()
    agents = Agent.query.all()

    if request.method == 'POST':
        lead.customer_name = request.form.get(
            'customer_name', lead.customer_name).strip()
        lead.customer_phone = request.form.get(
            'customer_phone', lead.customer_phone).strip()
        lead.customer_email = request.form.get(
            'customer_email', lead.customer_email or '').strip()
        lead.customer_city = request.form.get(
            'customer_city', lead.customer_city or '').strip()
        lead.source = request.form.get('source', lead.source)
        lead.stage = request.form.get('stage', lead.stage)
        lead.notes = request.form.get('notes', lead.notes or '').strip()
        budget = request.form.get('budget', type=float)
        if budget is not None:
            lead.budget = budget
        agent_id = request.form.get('agent_id', type=int)
        if agent_id:
            lead.agent_id = agent_id
        vehicle_id = request.form.get('vehicle_id', type=int)
        if vehicle_id:
            lead.vehicle_id = vehicle_id
        db.session.commit()
        log_admin_action(f"Updated lead {lead_id}", 'Leads')
        flash('Lead updated successfully!', 'success')
        return redirect(url_for('admin.all_leads'))

    return render_template('admin/edit_lead.html', lead=lead, dealers=dealers,
                           vehicles=vehicles, agents=agents, page='leads')


@admin_bp.route('/api/leads/<int:lead_id>/assign', methods=['POST'])
@require_permission_api('leads')
def assign_lead(lead_id):
    from models import Lead, Agent
    from extensions import db
    lead = Lead.query.get_or_404(lead_id)
    data = request.get_json(silent=True) or {}
    agent_id = data.get('agent_id')
    if agent_id:
        agent = Agent.query.get(agent_id)
        if agent:
            lead.agent_id = agent_id
            lead.assigned_to = agent.name
            db.session.commit()
            log_admin_action(
                f"Assigned lead {lead_id} to agent {agent.name}", 'Leads')
            return jsonify({'success': True, 'message': f'Lead assigned to {agent.name}'})
    return jsonify({'success': False, 'message': 'Invalid agent'})


@admin_bp.route('/api/leads/bulk-assign-dealer', methods=['POST'])
@require_permission_api('leads')
def bulk_assign_leads_to_dealer():
    """Assign multiple leads to a dealer at once."""
    from models import Lead, User
    from extensions import db
    data = request.get_json(silent=True) or {}
    lead_ids  = data.get('lead_ids', [])
    dealer_id = data.get('dealer_id')

    if not lead_ids or not dealer_id:
        return jsonify({'success': False, 'message': 'lead_ids and dealer_id are required'})

    dealer = User.query.filter_by(id=dealer_id, role='dealer').first()
    if not dealer:
        return jsonify({'success': False, 'message': 'Dealer not found'})

    updated = 0
    for lid in lead_ids:
        lead = Lead.query.get(lid)
        if lead:
            lead.dealer_id = dealer_id
            updated += 1

    db.session.commit()
    log_admin_action(f"Bulk-assigned {updated} leads to dealer {dealer.name}", 'Leads')
    return jsonify({'success': True, 'message': f'{updated} lead(s) assigned to {dealer.name}', 'count': updated})


@admin_bp.route('/leads/import', methods=['POST'])
@require_permission('leads')
def import_leads():
    """Import leads from CSV / XLSX / XLS directly into the leads table."""
    from models import Lead, User
    from extensions import db

    file = request.files.get('lead_file')
    if not file or not file.filename:
        return jsonify({'success': False, 'message': 'No file selected.'}), 400

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ('csv', 'xlsx', 'xls'):
        return jsonify({'success': False,
                        'message': f'Unsupported file format ".{ext}". Please upload CSV, XLSX, or XLS.'}), 400

    default_dealer_id = request.form.get('default_dealer_id', type=int)

    # ── Parse file into list of dicts ───────────────────────────────────────
    rows = []
    try:
        if ext == 'csv':
            content = file.stream.read().decode('utf-8-sig', errors='replace')
            reader = csv.DictReader(io.StringIO(content, newline=None))
            rows = [dict(r) for r in reader]
        elif ext == 'xlsx':
            import openpyxl
            wb = openpyxl.load_workbook(file.stream, read_only=True, data_only=True)
            ws = wb.active
            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c).strip() if c is not None else f'col_{j}' for j, c in enumerate(row)]
                else:
                    if all(c is None for c in row):
                        continue
                    rows.append(dict(zip(headers, [str(c).strip() if c is not None else '' for c in row])))
            wb.close()
        else:  # xls
            import pandas as pd
            df = pd.read_excel(file.stream, engine='xlrd', dtype=str)
            df = df.fillna('')
            rows = df.to_dict(orient='records')
    except ImportError as e:
        return jsonify({'success': False, 'message': f'Missing library: {e}. Run: pip install openpyxl xlrd pandas'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': f'Could not parse file: {e}'}), 400

    if not rows:
        return jsonify({'success': False, 'message': 'File is empty or has no data rows.'}), 400

    # ── Helper: flexible column lookup ──────────────────────────────────────
    def _get(row, *keys):
        for k in keys:
            for rk in row:
                if rk.strip().lower().replace(' ', '_') == k.lower().replace(' ', '_'):
                    v = str(row[rk]).strip()
                    if v and v.lower() not in ('none', 'nan', 'null', ''):
                        return v
        return ''

    # ── Process rows ────────────────────────────────────────────────────────
    imported = 0
    duplicates = 0
    failed = 0
    error_details = []
    leads_to_add = []

    for i, row in enumerate(rows, 1):
        name  = _get(row, 'customer_name', 'name', 'full_name', 'Customer Name', 'Name')
        phone = _get(row, 'customer_phone', 'phone', 'phone_number', 'mobile', 'Phone', 'Mobile')
        email = _get(row, 'customer_email', 'email', 'email_address', 'Email') or None
        city  = _get(row, 'customer_city', 'city', 'location', 'City') or None
        source = _get(row, 'source', 'lead_source', 'Source') or 'import'
        stage  = _get(row, 'stage', 'Stage') or 'new'
        notes  = _get(row, 'notes', 'Notes') or None
        budget_raw = _get(row, 'budget', 'Budget')

        if not name or not phone:
            failed += 1
            error_details.append(f'Row {i}: missing customer name or phone — skipped')
            continue

        # Duplicate check on phone and email
        dup_q = Lead.query.filter(Lead.customer_phone == phone)
        if email:
            from sqlalchemy import or_
            dup_q = Lead.query.filter(
                or_(Lead.customer_phone == phone, Lead.customer_email == email)
            )
        if dup_q.first():
            duplicates += 1
            continue

        # Dealer resolution
        dealer_id = None
        raw_did = _get(row, 'dealer_id', 'dealer', 'Dealer ID', 'Dealer')
        if raw_did.upper().startswith('D') and raw_did[1:].isdigit():
            matched = User.query.filter_by(display_id=raw_did.upper(), role='dealer').first()
            if matched:
                dealer_id = matched.id
        elif raw_did.isdigit():
            dealer_id = int(raw_did)
        dealer_id = dealer_id or default_dealer_id
        # No dealer? Import as unassigned (dealer_id stays None)

        budget = None
        if budget_raw:
            try:
                budget = float(budget_raw.replace(',', '').replace('₹', '').strip())
            except ValueError:
                pass

        leads_to_add.append(Lead(
            dealer_id=dealer_id,
            customer_name=name,
            customer_phone=phone,
            customer_email=email,
            customer_city=city,
            source=source,
            stage=stage,
            budget=budget,
            notes=notes,
        ))
        imported += 1

    if leads_to_add:
        db.session.bulk_save_objects(leads_to_add)
        db.session.commit()

    log_admin_action(f"Imported {imported} leads from {ext.upper()} file", 'Leads')

    return jsonify({
        'success': True,
        'total': len(rows),
        'imported': imported,
        'duplicates': duplicates,
        'failed': failed,
        'errors': error_details[:10],
    })

@admin_bp.route('/api/leads/<int:lead_id>/update-stage', methods=['POST'])
@require_permission_api('leads')
def update_lead_stage(lead_id):
    from models import Lead
    from extensions import db
    lead = Lead.query.get_or_404(lead_id)
    data = request.get_json(silent=True) or {}
    stage = data.get('stage', lead.stage)
    lead.stage = stage
    db.session.commit()
    log_admin_action(f"Updated lead {lead_id} stage to {stage}", 'Leads')
    return jsonify({'success': True, 'stage': stage})


@admin_bp.route('/api/leads/<int:lead_id>/delete', methods=['POST'])
@require_permission_api('leads')
def delete_lead(lead_id):
    from models import Lead
    from extensions import db
    lead = Lead.query.get_or_404(lead_id)
    db.session.delete(lead)
    db.session.commit()
    log_admin_action(f"Deleted lead {lead_id}", 'Leads')
    return jsonify({'success': True, 'message': 'Lead deleted.'})


# ─── Sales / Deals ────────────────────────────────────────────────────────────

@admin_bp.route('/sales')
@require_permission('leads')
def all_sales():
    from models import Deal
    deals = Deal.query.order_by(Deal.created_at.desc()).all()
    total_revenue = sum(
        d.final_price or 0 for d in deals if d.status == 'delivered')
    return render_template('admin/sales.html', deals=deals, total_revenue=total_revenue, page='sales')


# ─── Inquiries ────────────────────────────────────────────────────────────────

@admin_bp.route('/inquiries')
@require_permission('leads')
def all_inquiries():
    from models import Inquiry, Vehicle, User
    inquiries_raw = Inquiry.query.order_by(Inquiry.created_at.desc()).all()
    inquiries = []
    for inq in inquiries_raw:
        vehicle = Vehicle.query.get(inq.vehicle_id) if inq.vehicle_id else None
        dealer  = User.query.get(vehicle.dealer_id) if vehicle and vehicle.dealer_id else None
        inquiries.append({
            'id':           inq.id,
            'name':         inq.name,
            'phone':        inq.phone,
            'email':        inq.email,
            'message':      inq.message,
            'inquiry_type': inq.inquiry_type,
            'status':       inq.status,
            'created_at':   inq.created_at,
            'vehicle_name': f"{vehicle.year} {vehicle.make} {vehicle.model}" if vehicle else '—',
            'dealer_name':  dealer.name if dealer else '—',
            'dealer_company': dealer.company_name if dealer else '—',
        })
    return render_template('admin/inquiries.html', inquiries=inquiries, page='inquiries')


# ─── Reports ──────────────────────────────────────────────────────────────────

@admin_bp.route('/reports')
@require_permission('reports')
def reports():
    return render_template('admin/reports.html', page='reports')


@admin_bp.route('/api/reports/generate')
@require_permission_api('reports')
def generate_report():
    from models import User, Vehicle, Lead, Deal
    rtype = request.args.get('type', 'dealer').lower()
    fmt = request.args.get('format', 'csv').lower()
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'{rtype}_report_{timestamp}'

    if rtype == 'dealer':
        headers = ['Display ID', 'Name', 'Email', 'Phone', 'Company',
                   'City', 'GST', 'Plan', 'Active', 'Registered']
        rows = [[(d.display_id or ('D' + str(d.id))), d.name, d.email, d.phone or '', d.company_name or '',
                 d.city or '', d.gst_number or '', d.subscription_plan,
                 'Yes' if d.is_active else 'No',
                 d.created_at.strftime('%Y-%m-%d') if d.created_at else '']
                for d in User.query.filter_by(role='dealer').all()]
    elif rtype == 'vehicle':
        headers = ['ID', 'Make', 'Model', 'Year', 'Fuel', 'Price',
                   'Mileage', 'Status', 'Approval', 'Dealer', 'Date']
        rows = [[v.id, v.make, v.model, v.year or '', v.fuel_type or '',
                 v.price or 0, v.mileage or 0, v.status,
                 v.approval_status, v.dealer.name if v.dealer else '',
                 v.created_at.strftime('%Y-%m-%d') if v.created_at else '']
                for v in Vehicle.query.all()]
    elif rtype == 'lead':
        headers = ['ID', 'Customer', 'Phone',
                   'Vehicle', 'Stage', 'Source', 'Date']
        rows = [[l.id, l.customer_name or '', l.customer_phone or '',
                 f"{l.vehicle.make} {l.vehicle.model}" if l.vehicle else '',
                 l.stage, l.source,
                 l.created_at.strftime('%Y-%m-%d') if l.created_at else '']
                for l in Lead.query.all()]
    elif rtype in ('sales', 'revenue'):
        headers = ['ID', 'Vehicle', 'Dealer', 'Sale Price', 'Payment', 'Date']
        rows = [[d.id, f"{d.vehicle.make} {d.vehicle.model}" if d.vehicle else '',
                 d.dealer.name if d.dealer else '',
                 d.final_price or 0, d.payment_mode or '',
                 d.created_at.strftime('%Y-%m-%d') if d.created_at else '']
                for d in Deal.query.all()]
    else:
        headers = ['Metric', 'Count']
        rows = [
            ['Total Dealers', User.query.filter_by(role='dealer').count()],
            ['Total Users', User.query.filter_by(role='user').count()],
            ['Total Vehicles', Vehicle.query.count()],
            ['Total Leads', Lead.query.count()],
            ['Total Sales', Deal.query.count()],
        ]

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    log_admin_action(f"Generated {rtype} report ({fmt})", 'Reports')
    return response


# ─── Activity Logs ────────────────────────────────────────────────────────────

@admin_bp.route('/activity')
@require_permission('reports')
def activity():
    from models import AdminLog
    q = AdminLog.query

    # ── Filters ──────────────────────────────────────────────────────────────
    role_filter   = request.args.get('role', '').strip()
    module_filter = request.args.get('module', '').strip()
    date_filter   = request.args.get('date', '').strip()
    search        = request.args.get('search', '').strip()
    page          = max(1, int(request.args.get('page', 1) or 1))
    per_page      = 25

    if role_filter:
        q = q.filter(AdminLog.user_role == role_filter)
    if module_filter:
        q = q.filter(AdminLog.module == module_filter)
    if date_filter:
        try:
            from datetime import datetime as _dt
            date_obj = _dt.strptime(date_filter, '%Y-%m-%d')
            q = q.filter(
                AdminLog.created_at >= date_obj,
                AdminLog.created_at < _dt(date_obj.year, date_obj.month, date_obj.day, 23, 59, 59)
            )
        except ValueError:
            pass
    if search:
        like = f'%{search}%'
        q = q.filter(
            (AdminLog.action.ilike(like)) |
            (AdminLog.admin_user.ilike(like)) |
            (AdminLog.module.ilike(like)) |
            (AdminLog.ip_address.ilike(like))
        )

    # Keep the filtered-but-unordered query for aggregate stats — computing
    # COUNT(DISTINCT ...) against it (rather than the whole table) makes the
    # stat cards respect whatever role/module/date/search filter is active,
    # same as `total` does.
    base_q        = q
    total         = base_q.count()
    total_pages   = max(1, (total + per_page - 1) // per_page)
    page          = min(page, total_pages)
    logs          = base_q.order_by(AdminLog.created_at.desc()) \
                           .offset((page - 1) * per_page).limit(per_page).all()

    # Distinct modules for filter dropdown (also reused as "Pages/Modules Tracked")
    all_modules = [r[0] for r in AdminLog.query.with_entities(AdminLog.module).distinct().all() if r[0]]

    # ── Unique-actor aggregation (fixes inflated Unique Users / Roles Tracked) ──
    uniq_expr      = _activity_unique_key_expr()
    unique_users   = base_q.with_entities(func.count(func.distinct(uniq_expr))).scalar() or 0
    roles_tracked  = base_q.with_entities(func.count(func.distinct(AdminLog.user_role))).scalar() or 0
    modules_tracked = base_q.with_entities(func.count(func.distinct(AdminLog.module))).scalar() or 0

    from datetime import timedelta
    return render_template(
        'admin/activity.html',
        logs=logs, page_num=page, per_page=per_page,
        total=total, total_pages=total_pages,
        unique_users=unique_users, roles_tracked=roles_tracked,
        modules_tracked=modules_tracked,
        role_filter=role_filter, module_filter=module_filter,
        date_filter=date_filter, search=search,
        all_modules=sorted(all_modules),
        page='activity',
        timedelta=timedelta,
    )


@admin_bp.route('/activity/export')
@require_permission('reports')
def activity_export():
    from models import AdminLog
    import csv, io
    q = AdminLog.query

    role_filter   = request.args.get('role', '').strip()
    module_filter = request.args.get('module', '').strip()
    date_filter   = request.args.get('date', '').strip()
    search        = request.args.get('search', '').strip()

    if role_filter:
        q = q.filter(AdminLog.user_role == role_filter)
    if module_filter:
        q = q.filter(AdminLog.module == module_filter)
    if date_filter:
        try:
            from datetime import datetime as _dt
            date_obj = _dt.strptime(date_filter, '%Y-%m-%d')
            q = q.filter(
                AdminLog.created_at >= date_obj,
                AdminLog.created_at < _dt(date_obj.year, date_obj.month, date_obj.day, 23, 59, 59)
            )
        except ValueError:
            pass
    if search:
        like = f'%{search}%'
        q = q.filter(
            (AdminLog.action.ilike(like)) |
            (AdminLog.admin_user.ilike(like)) |
            (AdminLog.module.ilike(like)) |
            (AdminLog.ip_address.ilike(like))
        )

    logs = q.order_by(AdminLog.created_at.desc()).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['#', 'User Role', 'User Name', 'Action', 'Description', 'Module',
                      'Date & Time', 'Timezone', 'IP Address', 'Device', 'Browser', 'Status'])
    for i, log in enumerate(logs, 1):
        writer.writerow([
            i,
            getattr(log, 'user_role', 'Admin'),
            log.admin_user,
            log.action,
            getattr(log, 'description', None) or log.action,
            log.module,
            log.created_at.strftime('%d %b %Y, %I:%M %p') if log.created_at else '',
            getattr(log, 'timezone', None) or 'Asia/Kolkata (IST)',
            log.ip_address,
            getattr(log, 'device', None) or '',
            getattr(log, 'browser', None) or '',
            getattr(log, 'status', 'Success'),
        ])
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment;filename=activity_logs.csv'}
    )


# ─── Visitor Logs ─────────────────────────────────────────────────────────────

@admin_bp.route('/visitor-logs')
@require_permission('reports')
def visitor_logs():
    from models import VisitorLog
    search     = request.args.get('search', '').strip()
    date_f     = request.args.get('date', '').strip()
    device_f   = request.args.get('device', '').strip()
    page       = max(1, int(request.args.get('page', 1) or 1))
    per_page   = 30

    q = VisitorLog.query
    if search:
        like = f'%{search}%'
        q = q.filter(
            VisitorLog.ip_address.ilike(like) |
            VisitorLog.browser.ilike(like) |
            VisitorLog.operating_system.ilike(like) |
            VisitorLog.page_url.ilike(like) |
            VisitorLog.country.ilike(like) |
            VisitorLog.visitor_name.ilike(like)
        )
    if date_f:
        try:
            from datetime import datetime as _dt2
            d = _dt2.strptime(date_f, '%Y-%m-%d')
            q = q.filter(
                VisitorLog.visited_at >= d,
                VisitorLog.visited_at < _dt2(d.year, d.month, d.day, 23, 59, 59)
            )
        except ValueError:
            pass
    if device_f:
        q = q.filter(VisitorLog.device_type == device_f)

    total       = q.count()
    total_pages = max(1, (total + per_page - 1) // per_page)
    page        = min(page, total_pages)
    logs        = q.order_by(VisitorLog.visited_at.desc()).offset((page - 1) * per_page).limit(per_page).all()

    export = request.args.get('export', '')
    if export == 'csv':
        all_logs = q.order_by(VisitorLog.visited_at.desc()).all()
        import csv as _csv, io as _io
        buf = _io.StringIO()
        w = _csv.writer(buf)
        w.writerow(['IP Address', 'Visitor Name', 'Role', 'Country', 'City', 'Browser', 'OS', 'Device', 'Page URL', 'Referrer', 'Visited At'])
        for v in all_logs:
            w.writerow([
                v.ip_address, v.visitor_name or 'Guest', v.visitor_role or '',
                v.country or '', v.city or '',
                v.browser or '', v.operating_system or '', v.device_type or '',
                v.page_url or '', v.referrer or '',
                v.visited_at.strftime('%d %b %Y %H:%M:%S') if v.visited_at else ''
            ])
        return Response(buf.getvalue(), mimetype='text/csv',
                        headers={'Content-Disposition': 'attachment;filename=visitor_logs.csv'})

    uniq_expr = _visitor_unique_key_expr()
    stats = {
        # Total Visits stays a raw page-view count (unchanged behaviour).
        'total':   VisitorLog.query.count(),
        # Total Users = unique visitors (logged-in user_id, else ip+device+session).
        'users':   VisitorLog.query.with_entities(func.count(func.distinct(uniq_expr))).scalar() or 0,
        # Desktop / Mobile / Tablet now count UNIQUE visitors per device,
        # not raw page hits, so one person browsing several pages on the
        # same device is counted once instead of once per page.
        'desktop': VisitorLog.query.filter(VisitorLog.device_type == 'Desktop')
                                    .with_entities(func.count(func.distinct(uniq_expr))).scalar() or 0,
        'mobile':  VisitorLog.query.filter(VisitorLog.device_type == 'Mobile')
                                    .with_entities(func.count(func.distinct(uniq_expr))).scalar() or 0,
        'tablet':  VisitorLog.query.filter(VisitorLog.device_type == 'Tablet')
                                    .with_entities(func.count(func.distinct(uniq_expr))).scalar() or 0,
    }
    return render_template('admin/visitor_logs.html',
                           logs=logs, page_num=page, total=total,
                           total_pages=total_pages, per_page=per_page,
                           search=search, date_f=date_f, device_f=device_f,
                           stats=stats, page='visitor_logs')


# ─── Real-Time Polling API: Activity Logs ─────────────────────────────────────

@admin_bp.route('/api/activity-poll')
@admin_api_login_required
def api_activity_poll():
    """
    Lightweight polling endpoint for Activity Logs real-time updates.
    Returns new AdminLog rows with id > since_id, preserving active filters.
    Called by the activity.html page every 30 seconds via fetch().
    """
    from models import AdminLog
    since_id      = request.args.get('since_id', 0, type=int)
    role_filter   = request.args.get('role', '').strip()
    module_filter = request.args.get('module', '').strip()

    q = AdminLog.query.filter(AdminLog.id > since_id)
    if role_filter:
        q = q.filter(AdminLog.user_role == role_filter)
    if module_filter:
        q = q.filter(AdminLog.module == module_filter)

    new_logs = q.order_by(AdminLog.id.desc()).limit(50).all()
    total    = AdminLog.query.count()

    # Same filtered base (minus since_id, which only limits *new* rows) used
    # for the unique-actor stat cards, so they stay in sync with the page.
    stats_q = AdminLog.query
    if role_filter:
        stats_q = stats_q.filter(AdminLog.user_role == role_filter)
    if module_filter:
        stats_q = stats_q.filter(AdminLog.module == module_filter)
    uniq_expr = _activity_unique_key_expr()
    stats = {
        'total':           total,
        'unique_users':    stats_q.with_entities(func.count(func.distinct(uniq_expr))).scalar() or 0,
        'roles_tracked':   stats_q.with_entities(func.count(func.distinct(AdminLog.user_role))).scalar() or 0,
        'modules_tracked': stats_q.with_entities(func.count(func.distinct(AdminLog.module))).scalar() or 0,
    }

    rows = []
    for log in new_logs:
        rows.append({
            'id':         log.id,
            'user_role':  log.user_role or 'Admin',
            'admin_user': log.admin_user or '',
            'action':     log.action or '',
            'module':     log.module or 'System',
            'ip_address': log.ip_address or '—',
            'device':     getattr(log, 'device', None) or '',
            'browser':    getattr(log, 'browser', None) or '',
            'status':     log.status or 'Success',
            # Stored value is naive IST (UTC+5:30) — tag it with the correct
            # offset (not "Z"/UTC) so the browser's Date parsing is accurate
            # regardless of the viewer's own timezone.
            'created_at': log.created_at.strftime('%Y-%m-%dT%H:%M:%S') + '+05:30'
                          if log.created_at else None,
        })

    return jsonify({'success': True, 'rows': rows, 'total': total, 'stats': stats})


# ─── Real-Time Polling API: Visitor Logs ──────────────────────────────────────

@admin_bp.route('/api/visitor-poll')
@admin_api_login_required
def api_visitor_poll():
    """
    Lightweight polling endpoint for Visitor Logs real-time updates.
    Returns new VisitorLog rows with id > since_id.
    Called by the visitor_logs.html page every 30 seconds via fetch().
    """
    from models import VisitorLog
    since_id = request.args.get('since_id', 0, type=int)
    device_f = request.args.get('device', '').strip()

    q = VisitorLog.query.filter(VisitorLog.id > since_id)
    if device_f:
        q = q.filter(VisitorLog.device_type == device_f)

    new_logs = q.order_by(VisitorLog.id.desc()).limit(50).all()

    uniq_expr = _visitor_unique_key_expr()
    stats = {
        'total':   VisitorLog.query.count(),
        'users':   VisitorLog.query.with_entities(func.count(func.distinct(uniq_expr))).scalar() or 0,
        'desktop': VisitorLog.query.filter(VisitorLog.device_type == 'Desktop')
                                    .with_entities(func.count(func.distinct(uniq_expr))).scalar() or 0,
        'mobile':  VisitorLog.query.filter(VisitorLog.device_type == 'Mobile')
                                    .with_entities(func.count(func.distinct(uniq_expr))).scalar() or 0,
        'tablet':  VisitorLog.query.filter(VisitorLog.device_type == 'Tablet')
                                    .with_entities(func.count(func.distinct(uniq_expr))).scalar() or 0,
    }

    rows = []
    for v in new_logs:
        rows.append({
            'id':               v.id,
            'ip_address':       v.ip_address or '—',
            'visitor_name':     v.visitor_name or '',
            'visitor_role':     v.visitor_role or '',
            'country':          v.country or '',
            'city':             v.city or '',
            'browser':          v.browser or '—',
            'operating_system': v.operating_system or '',
            'device_type':      v.device_type or '—',
            'page_url':         v.page_url or '—',
            'referrer':         v.referrer or '',
            # Stored value is naive IST (UTC+5:30) — tag with the correct
            # offset so the browser parses/localises it accurately.
            'visited_at':       v.visited_at.strftime('%Y-%m-%dT%H:%M:%S') + '+05:30'
                                if v.visited_at else None,
        })

    return jsonify({'success': True, 'rows': rows, 'stats': stats})


@admin_bp.route('/settings')
@require_permission('settings')
def settings():
    return render_template('admin/settings.html', page='settings',
                           gs=GENERAL_SETTINGS, es=EMAIL_SETTINGS)


@admin_bp.route('/settings/save', methods=['POST'])
@require_permission('settings')
def save_settings():
    global GENERAL_SETTINGS
    for key in GENERAL_SETTINGS:
        val = request.form.get(key, '').strip()
        if val:
            GENERAL_SETTINGS[key] = val
    log_admin_action('General settings updated', 'Settings')
    flash('Settings saved successfully!', 'success')
    return redirect(url_for('admin.settings'))


@admin_bp.route('/settings/change-password', methods=['POST'])
@require_permission_api('settings')
def change_password():
    global ADMIN_CREDS
    data = request.get_json(silent=True) or {}
    current = data.get('current', '')
    new_pw = data.get('newPassword', '')
    if current != ADMIN_CREDS['password']:
        return jsonify({'success': False, 'message': 'Current password is incorrect.'})
    if len(new_pw) < 6:
        return jsonify({'success': False, 'message': 'Password must be at least 6 characters.'})
    ADMIN_CREDS['password'] = new_pw
    log_admin_action('Admin changed password', 'Settings')
    return jsonify({'success': True, 'message': 'Password updated successfully!'})


# ─── API: Chart data ──────────────────────────────────────────────────────────

@admin_bp.route('/api/chart-data')
@admin_api_login_required
def chart_data():
    from models import Deal, Lead, User
    from sqlalchemy import extract
    from extensions import db

    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    current_year = _now_ist().year

    sales_by_month = []
    revenue_by_month = []
    for m in range(1, 7):
        deals = Deal.query.filter(
            extract('month', Deal.created_at) == m,
            extract('year',  Deal.created_at) == current_year,
        ).all()
        sales_by_month.append(len(deals))
        revenue_by_month.append(sum(d.final_price or 0 for d in deals))

    leads_by_month = []
    for m in range(1, 7):
        count = Lead.query.filter(
            extract('month', Lead.created_at) == m,
            extract('year',  Lead.created_at) == current_year
        ).count()
        leads_by_month.append(count)

    dealers = User.query.filter_by(role='dealer').all()
    dealer_names = [d.company_name or d.name for d in dealers[:5]]
    dealer_sales = [len(d.vehicles) for d in dealers[:5]]

    return jsonify({
        'months':       months[:6],
        'sales':        sales_by_month,
        'revenue':      revenue_by_month,
        'leads':        leads_by_month,
        'dealers':      dealer_names,
        'dealer_sales': dealer_sales,
    })


# ─── Notifications ────────────────────────────────────────────────────────────

@admin_bp.route('/notifications')
@admin_login_required
def notifications():
    from models import AdminLog
    logs = AdminLog.query.order_by(AdminLog.created_at.desc()).limit(20).all()
    return render_template('admin/notifications.html', logs=logs, page='notifications')


# ═══════════════════════════════════════════════════════════════════════════════
# KYC VERIFICATION MODULE
# ═══════════════════════════════════════════════════════════════════════════════

def _kyc_upload_folder():
    return current_app.config.get(
        'KYC_UPLOAD_FOLDER',
        os.path.join(os.path.dirname(__file__), '..',
                     'static', 'uploads', 'dealers')
    )


def _vehicle_upload_folder():
    # Must match UPLOAD_FOLDER (static/images/uploads) so filenames stored in
    # the DB resolve correctly in all templates that use /static/images/uploads/{filename}
    return current_app.config.get(
        'UPLOAD_FOLDER',
        os.path.join(os.path.dirname(__file__), '..',
                     'static', 'images', 'uploads')
    )


def _kyc_image_upload_dir():
    """Absolute path to the flat KYC image folder (for single document_image uploads)."""
    folder = os.path.join(
        current_app.root_path,
        current_app.config.get('KYC_IMAGE_FOLDER', 'static/uploads/kyc')
    )
    os.makedirs(folder, exist_ok=True)
    return folder


def _allowed_kyc_file(filename):
    return (
        '.' in filename
        and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
    )


@admin_bp.route('/kyc')
@require_permission('kyc')
def kyc_list():
    """Admin KYC dashboard - list all dealers with their KYC status."""
    from models import User, DealerKYC
    dealers = User.query.filter_by(role='dealer').order_by(
        User.created_at.desc()).all()
    kyc_map = {k.dealer_id: k for k in DealerKYC.query.all()}
    pending_count = sum(1 for d in dealers if kyc_map.get(
        d.id) and kyc_map[d.id].kyc_status == 'pending')
    approved_count = sum(1 for d in dealers if kyc_map.get(
        d.id) and kyc_map[d.id].kyc_status == 'approved')
    rejected_count = sum(1 for d in dealers if kyc_map.get(
        d.id) and kyc_map[d.id].kyc_status == 'rejected')
    no_kyc_count = sum(1 for d in dealers if not kyc_map.get(d.id))
    return render_template('admin/kyc_list.html',
                           dealers=dealers, kyc_map=kyc_map,
                           pending_count=pending_count, approved_count=approved_count,
                           rejected_count=rejected_count, no_kyc_count=no_kyc_count,
                           page='kyc')


@admin_bp.route('/kyc/<int:dealer_id>')
@require_permission('kyc')
def kyc_detail(dealer_id):
    """View KYC documents for a specific dealer."""
    from models import User, DealerKYC
    dealer = User.query.get_or_404(dealer_id)
    kyc = DealerKYC.query.filter_by(dealer_id=dealer_id).first()
    return render_template('admin/kyc_detail.html', dealer=dealer, kyc=kyc, page='kyc')


@admin_bp.route('/kyc/<int:dealer_id>/upload', methods=['POST'])
@require_permission_api('kyc')
def kyc_upload_docs(dealer_id):
    """Admin can upload/replace KYC documents (aadhaar/pan) on behalf of a dealer."""
    from models import User, DealerKYC
    from extensions import db
    from utils.upload_helpers import save_image, delete_image, validate_image

    dealer = User.query.get_or_404(dealer_id)
    kyc = DealerKYC.query.filter_by(dealer_id=dealer_id).first()
    if not kyc:
        kyc = DealerKYC(dealer_id=dealer_id, kyc_status='pending')
        db.session.add(kyc)

    folder = os.path.join(_kyc_upload_folder(), str(dealer_id))
    os.makedirs(folder, exist_ok=True)
    errors = []

    for doc_key in ('aadhaar_front', 'aadhaar_back', 'pan_card'):
        f = request.files.get(doc_key)
        if not f or not f.filename:
            continue
        ok, err = validate_image(f)
        if not ok:
            errors.append(f'{doc_key}: {err}')
            continue
        old = getattr(kyc, doc_key)
        if old:
            delete_image(folder, old)
        saved = save_image(f, folder, prefix=doc_key.replace('_', '-'), vehicle_mode=False)
        if saved:
            setattr(kyc, doc_key, saved)
        else:
            errors.append(f'Failed to save {doc_key}.')

    if errors:
        for e in errors:
            flash(e, 'error')
    else:
        flash('KYC documents uploaded successfully.', 'success')

    db.session.commit()
    log_admin_action(f'Uploaded KYC docs for dealer {dealer.name}', 'KYC')
    return redirect(url_for('admin.kyc_detail', dealer_id=dealer_id))


# ── NEW: Upload / replace single KYC document image ──────────────────────────

@admin_bp.route('/kyc/<int:dealer_id>/upload-image', methods=['POST'])
@require_permission_api('kyc')
def kyc_upload_image(dealer_id):
    """Admin: upload or replace the single KYC document image for a dealer."""
    from models import DealerKYC
    from extensions import db

    kyc = DealerKYC.query.filter_by(dealer_id=dealer_id).first()
    if not kyc:
        kyc = DealerKYC(dealer_id=dealer_id, kyc_status='pending')
        db.session.add(kyc)
        db.session.flush()

    if 'kyc_image' not in request.files:
        flash('No file part in the request.', 'error')
        return redirect(url_for('admin.kyc_detail', dealer_id=dealer_id))

    file = request.files['kyc_image']

    if file.filename == '':
        flash('No file selected.', 'error')
        return redirect(url_for('admin.kyc_detail', dealer_id=dealer_id))

    if not _allowed_kyc_file(file.filename):
        flash('Invalid file type. Allowed: JPG, PNG, GIF, WEBP, PDF.', 'error')
        return redirect(url_for('admin.kyc_detail', dealer_id=dealer_id))

    # Size check
    file.seek(0, os.SEEK_END)
    size = file.tell()
    file.seek(0)
    if size > MAX_KYC_IMAGE_SIZE:
        flash('File too large. Maximum allowed size is 5 MB.', 'error')
        return redirect(url_for('admin.kyc_detail', dealer_id=dealer_id))

    upload_dir = _kyc_image_upload_dir()

    # Delete old file from disk if present
    if kyc.document_image:
        old_path = os.path.join(upload_dir, kyc.document_image)
        if os.path.isfile(old_path):
            try:
                os.remove(old_path)
            except OSError:
                pass

    # Save with a unique safe filename
    ext = secure_filename(file.filename).rsplit('.', 1)[-1].lower()
    new_filename = f"kyc_{dealer_id}_{uuid.uuid4().hex}.{ext}"
    save_path = os.path.join(upload_dir, new_filename)

    try:
        file.save(save_path)
    except Exception as e:
        current_app.logger.error(f"KYC image save failed: {e}")
        flash('Failed to save file. Please try again.', 'error')
        return redirect(url_for('admin.kyc_detail', dealer_id=dealer_id))

    kyc.document_image = new_filename
    db.session.commit()
    log_admin_action(f'Uploaded KYC image for dealer_id={dealer_id}', 'KYC')
    flash('KYC image uploaded successfully.', 'success')
    return redirect(url_for('admin.kyc_detail', dealer_id=dealer_id))


# ── NEW: Delete single KYC document image ────────────────────────────────────

@admin_bp.route('/kyc/<int:dealer_id>/delete-image', methods=['POST'])
@require_permission_api('kyc')
def kyc_delete_image(dealer_id):
    """Admin: remove the KYC document image for a dealer."""
    from models import DealerKYC
    from extensions import db

    kyc = DealerKYC.query.filter_by(dealer_id=dealer_id).first()

    if not kyc or not kyc.document_image:
        flash('No image to remove.', 'warning')
        return redirect(url_for('admin.kyc_detail', dealer_id=dealer_id))

    upload_dir = _kyc_image_upload_dir()
    file_path = os.path.join(upload_dir, kyc.document_image)

    if os.path.isfile(file_path):
        try:
            os.remove(file_path)
        except OSError as e:
            current_app.logger.error(f"KYC image delete failed: {e}")
            flash(
                'Could not delete file from disk, but record has been cleared.', 'warning')

    kyc.document_image = None
    db.session.commit()
    log_admin_action(f'Deleted KYC image for dealer_id={dealer_id}', 'KYC')
    flash('KYC image removed successfully.', 'success')
    return redirect(url_for('admin.kyc_detail', dealer_id=dealer_id))


# ─── KYC Approve / Reject / Reset API ────────────────────────────────────────
# Helper: return live counter snapshot for dashboard cards

def _kyc_counts():
    """Return {pending, approved, rejected, none} counts for all dealers."""
    from models import User, DealerKYC
    from sqlalchemy import func
    dealers_total = User.query.filter_by(role='dealer').count()
    kyc_rows = DealerKYC.query.all()
    counts = {'approved': 0, 'rejected': 0, 'pending': 0, 'none': 0}
    kyc_dealer_ids = set()
    for k in kyc_rows:
        kyc_dealer_ids.add(k.dealer_id)
        s = k.kyc_status or 'pending'
        if s in counts:
            counts[s] += 1
        else:
            counts['pending'] += 1
    # dealers with no KYC record
    all_dealer_ids = {u.id for u in User.query.filter_by(role='dealer').all()}
    counts['none'] = len(all_dealer_ids - kyc_dealer_ids)
    return counts


# Helper: write a KYCReview audit record
def _log_kyc_review(dealer_id, document_type, status, reason, previous_status, reviewer):
    from models import KYCReview
    from extensions import db
    record = KYCReview(
        dealer_id=dealer_id,
        document_type=document_type,
        status=status,
        reason=reason or None,
        previous_status=previous_status,
        reviewed_by=reviewer,
    )
    db.session.add(record)


@admin_bp.route('/api/kyc/<int:dealer_id>/approve', methods=['POST'])
@require_permission_api('kyc')
def api_approve_kyc(dealer_id):
    from models import User, DealerKYC, DealerNotification
    from extensions import db
    dealer = User.query.get_or_404(dealer_id)
    kyc = DealerKYC.query.filter_by(dealer_id=dealer_id).first()
    if not kyc:
        kyc = DealerKYC(dealer_id=dealer_id)
        db.session.add(kyc)
    reviewer = session.get('admin_username', 'admin')
    now = _now_ist()
    for doc in ('aadhaar_front', 'aadhaar_back', 'pan_card'):
        prev = getattr(kyc, doc + '_status') or 'pending'
        setattr(kyc, doc + '_status', 'approved')
        setattr(kyc, doc + '_reject', None)
        setattr(kyc, doc + '_reviewed_by', reviewer)
        setattr(kyc, doc + '_reviewed_at', now)
        _log_kyc_review(dealer_id, doc, 'approved', None, prev, reviewer)
    prev_overall = kyc.kyc_status or 'pending'
    kyc.kyc_status = 'approved'
    kyc.rejection_reason = None
    kyc.reviewed_at = now
    kyc.reviewed_by = reviewer
    dealer.is_active = True
    _log_kyc_review(dealer_id, 'complete_kyc', 'approved', None, prev_overall, reviewer)
    db.session.add(DealerNotification(
        dealer_id=dealer_id,
        title='KYC Approved — Account Activated',
        message='Your KYC verification has been approved. Your dealer account is now fully active and all features are unlocked.',
        notif_type='success'
    ))
    db.session.commit()
    log_admin_action(f'KYC approved for dealer {dealer.name}', 'KYC')
    return jsonify({'success': True, 'message': f'KYC approved for {dealer.name}',
                    'counts': _kyc_counts()})


@admin_bp.route('/api/kyc/<int:dealer_id>/reject', methods=['POST'])
@require_permission_api('kyc')
def api_reject_kyc(dealer_id):
    from models import User, DealerKYC, DealerNotification
    from extensions import db
    dealer = User.query.get_or_404(dealer_id)
    data = request.get_json(silent=True) or {}
    reason = data.get('reason', '').strip()
    if not reason:
        return jsonify({'success': False, 'message': 'Rejection reason is required.'})
    kyc = DealerKYC.query.filter_by(dealer_id=dealer_id).first()
    if not kyc:
        kyc = DealerKYC(dealer_id=dealer_id)
        db.session.add(kyc)
    reviewer = session.get('admin_username', 'admin')
    now = _now_ist()
    prev_overall = kyc.kyc_status or 'pending'
    kyc.kyc_status = 'rejected'
    kyc.rejection_reason = reason
    kyc.reviewed_at = now
    kyc.reviewed_by = reviewer
    for doc in ('aadhaar_front', 'aadhaar_back', 'pan_card'):
        prev = getattr(kyc, doc + '_status') or 'pending'
        if prev == 'pending':
            setattr(kyc, doc + '_status', 'rejected')
            setattr(kyc, doc + '_reject', reason)
            setattr(kyc, doc + '_reviewed_by', reviewer)
            setattr(kyc, doc + '_reviewed_at', now)
            _log_kyc_review(dealer_id, doc, 'rejected', reason, prev, reviewer)
    _log_kyc_review(dealer_id, 'complete_kyc', 'rejected', reason, prev_overall, reviewer)
    db.session.add(DealerNotification(
        dealer_id=dealer_id,
        title='KYC Application Rejected',
        message=f'Your KYC application has been rejected. Reason: {reason}. Please re-upload your documents.',
        notif_type='danger'
    ))
    db.session.commit()
    log_admin_action(f'KYC rejected for dealer {dealer.name}: {reason}', 'KYC')
    return jsonify({'success': True, 'message': f'KYC rejected for {dealer.name}',
                    'counts': _kyc_counts()})


@admin_bp.route('/api/kyc/<int:dealer_id>/reset', methods=['POST'])
@require_permission_api('kyc')
def api_reset_kyc(dealer_id):
    from models import DealerKYC
    from extensions import db
    kyc = DealerKYC.query.filter_by(dealer_id=dealer_id).first()
    reviewer = session.get('admin_username', 'admin')
    if kyc:
        prev_overall = kyc.kyc_status or 'pending'
        kyc.kyc_status = 'pending'
        kyc.rejection_reason = None
        kyc.reviewed_at = None
        kyc.reviewed_by = None
        for doc in ('aadhaar_front', 'aadhaar_back', 'pan_card'):
            setattr(kyc, doc + '_status', 'pending')
            setattr(kyc, doc + '_reject', None)
        _log_kyc_review(dealer_id, 'complete_kyc', 'reset', None, prev_overall, reviewer)
        db.session.commit()
    log_admin_action(f'KYC reset to pending for dealer_id={dealer_id}', 'KYC')
    return jsonify({'success': True, 'counts': _kyc_counts()})


# ═══════════════════════════════════════════════════════════════════════════════
# VEHICLE IMAGE MANAGEMENT
# ═══════════════════════════════════════════════════════════════════════════════

@admin_bp.route('/vehicles/<int:vehicle_id>/images')
@require_permission('vehicles')
def vehicle_images(vehicle_id):
    from models import Vehicle, VehicleImage
    vehicle = Vehicle.query.get_or_404(vehicle_id)
    images = VehicleImage.query.filter_by(
        vehicle_id=vehicle_id).order_by(VehicleImage.sort_order).all()
    return render_template('admin/vehicle_images.html', vehicle=vehicle, images=images, page='vehicles')


@admin_bp.route('/api/vehicles/<int:vehicle_id>/images/upload', methods=['POST'])
@require_permission_api('vehicles')
def api_upload_vehicle_images(vehicle_id):
    from models import Vehicle, VehicleImage
    from extensions import db
    from utils.upload_helpers import save_image, validate_image

    Vehicle.query.get_or_404(vehicle_id)
    files = request.files.getlist('images')
    if not files:
        return jsonify({'success': False, 'message': 'No files uploaded.'})

    folder = _vehicle_upload_folder()
    os.makedirs(folder, exist_ok=True)

    saved_list = []
    errors = []
    existing = VehicleImage.query.filter_by(vehicle_id=vehicle_id).count()

    for f in files:
        if not f or not f.filename:
            continue
        ok, err = validate_image(f)
        if not ok:
            errors.append(f'{f.filename}: {err}')
            continue
        filename = save_image(f, folder, prefix='vimg')
        if filename:
            vi = VehicleImage(vehicle_id=vehicle_id,
                              filename=filename, sort_order=existing)
            db.session.add(vi)
            db.session.flush()
            v = Vehicle.query.get(vehicle_id)
            if not v.image_filename or v.image_filename == 'None':
                v.image_filename = filename
            db.session.commit()
            saved_list.append({'id': vi.id, 'filename': filename,
                               'url': f'/static/images/uploads/{filename}',
                               'sort_order': vi.sort_order})
            existing += 1
        else:
            errors.append(f'Failed to process {f.filename}.')

    if not saved_list and errors:
        return jsonify({'success': False, 'message': '; '.join(errors)})

    log_admin_action(
        f'Uploaded {len(saved_list)} images for vehicle {vehicle_id}', 'Vehicles')
    return jsonify({'success': True, 'uploaded': saved_list, 'errors': errors})


@admin_bp.route('/api/vehicles/images/<int:image_id>/delete', methods=['POST'])
@require_permission_api('vehicles')
def api_delete_vehicle_image(image_id):
    from models import VehicleImage, Vehicle
    from extensions import db
    from utils.upload_helpers import delete_image

    img = VehicleImage.query.get_or_404(image_id)
    vehicle_id = img.vehicle_id
    folder = _vehicle_upload_folder()
    delete_image(folder, img.filename)

    v = Vehicle.query.get(vehicle_id)
    if v and v.image_filename == img.filename:
        remaining = VehicleImage.query.filter(
            VehicleImage.vehicle_id == vehicle_id,
            VehicleImage.id != image_id
        ).order_by(VehicleImage.sort_order).first()
        v.image_filename = remaining.filename if remaining else None

    db.session.delete(img)
    db.session.commit()
    log_admin_action(f'Deleted vehicle image {image_id}', 'Vehicles')
    return jsonify({'success': True})


@admin_bp.route('/api/vehicles/images/<int:image_id>/set-primary', methods=['POST'])
@require_permission_api('vehicles')
def api_set_primary_vehicle_image(image_id):
    from models import VehicleImage, Vehicle
    from extensions import db
    img = VehicleImage.query.get_or_404(image_id)
    v = Vehicle.query.get(img.vehicle_id)
    if v:
        v.image_filename = img.filename
        db.session.commit()
    return jsonify({'success': True})


@admin_bp.route('/api/vehicles/images/reorder', methods=['POST'])
@require_permission_api('vehicles')
def api_reorder_vehicle_images():
    from models import VehicleImage
    from extensions import db
    data = request.get_json(silent=True) or {}
    order = data.get('order', [])
    for idx, image_id in enumerate(order):
        img = VehicleImage.query.get(image_id)
        if img:
            img.sort_order = idx
    db.session.commit()
    return jsonify({'success': True})


# ══════════════════════════════════════════════════════════════════════════════
# VEHICLE IMAGE EDITOR APIs  (Blur BG + Number Plate Hide + Save)
# Registered at /admin/api/vehicles/images/<image_id>/...
# Used by mask_editor_modal.html
# ══════════════════════════════════════════════════════════════════════════════

@admin_bp.route('/api/vehicles/images/<int:image_id>/detect-plate')
@require_permission_api('vehicles')
def api_vehicle_detect_plate(image_id):
    """Auto-detect number plate bounding box on a vehicle image."""
    from models import VehicleImage
    from background.utils import detect_number_plate
    from PIL import Image as _PILImg

    vi = VehicleImage.query.get_or_404(image_id)
    path = os.path.join(current_app.root_path, 'static', 'images', 'uploads', vi.filename)
    if not os.path.exists(path):
        return jsonify({'detected': False, 'message': 'File not found on disk'}), 404
    try:
        iw, ih = _PILImg.open(path).size
    except Exception as e:
        return jsonify({'detected': False, 'message': str(e)}), 400
    plate = detect_number_plate(path)
    if plate:
        x, y, w, h = plate
        return jsonify({'detected': True, 'x': x, 'y': y, 'width': w, 'height': h,
                        'img_width': iw, 'img_height': ih})
    return jsonify({'detected': False, 'img_width': iw, 'img_height': ih,
                    'message': 'Auto-detection failed. Use manual selection.'})


@admin_bp.route('/api/vehicles/images/<int:image_id>/blur-bg', methods=['POST'])
@require_permission_api('vehicles')
def api_vehicle_blur_bg(image_id):
    """
    AI background removal + depth-of-field blur applied to vehicle image.
    Result saved as temp processed file; call save-processed to commit.
    """
    from models import VehicleImage
    from PIL import Image as _PILImg
    from background.utils import (
        remove_bg_ai, apply_60_percent_background_blur,
        keep_largest_component, remove_persons_and_objects,
        remove_connected_persons, trim_side_cars, trim_top_objects,
        remove_thin_protrusions, restore_tyres, restore_windshield
    )

    vi = VehicleImage.query.get_or_404(image_id)
    src_path = os.path.join(current_app.root_path, 'static', 'images', 'uploads', vi.filename)
    if not os.path.exists(src_path):
        return jsonify({'success': False, 'error': 'Source file not found'}), 404

    result, method = remove_bg_ai(src_path, quality='standard')
    if result is None:
        try:
            result = _PILImg.open(src_path).convert('RGBA')
            method = 'original_kept'
        except Exception:
            return jsonify({'success': False, 'error': 'BG removal failed'}), 500

    try:
        result = keep_largest_component(result)
        result = remove_persons_and_objects(result)
        result = remove_connected_persons(result)
        result = trim_side_cars(result)
        result = trim_top_objects(result)
        result = remove_thin_protrusions(result)
        result = restore_tyres(result)
        result = restore_windshield(result)
    except Exception as _e:
        current_app.logger.warning(f'[admin blur_bg] mask cleanup (non-fatal): {_e}')

    pf = os.path.join(current_app.root_path, 'static', 'processed')
    os.makedirs(pf, exist_ok=True)
    out_fname = f'vblur_{image_id}_{uuid.uuid4().hex[:8]}.jpg'
    out_path  = os.path.join(pf, out_fname)
    try:
        blurred = apply_60_percent_background_blur(src_path, result)
        blurred = blurred.convert('RGB')
        # Apply tiled Caryanams watermark
        try:
            from PIL import ImageDraw, ImageFont
            wm_layer = _PILImg.new('RGBA', blurred.size, (0, 0, 0, 0))
            draw = ImageDraw.Draw(wm_layer)
            W, H = blurred.size
            font_paths_bold = [
                '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
                '/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf',
                'C:/Windows/Fonts/arialbd.ttf',
            ]
            font_paths_reg = [
                '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
                'C:/Windows/Fonts/arial.ttf',
            ]
            def _load_font(paths, size):
                for fp in paths:
                    try: return ImageFont.truetype(fp, size)
                    except: continue
                return ImageFont.load_default()
            sz1 = max(26, int(min(W, H) * 0.05))
            sz2 = max(14, int(min(W, H) * 0.025))
            font1 = _load_font(font_paths_bold, sz1)
            font2 = _load_font(font_paths_reg, sz2)
            color1 = (160, 160, 160, 100)
            color2 = (180, 140, 60, 120)
            _tmp = ImageDraw.Draw(_PILImg.new('RGBA', (1, 1)))
            def tsz(f, t):
                bb = _tmp.textbbox((0, 0), t, font=f)
                return bb[2]-bb[0], bb[3]-bb[1]
            tw1, th1 = tsz(font1, 'Caryanams')
            tw2, th2 = tsz(font2, 'Driven by Trust')
            tile_w = max(tw1, tw2) + int(W * 0.12)
            tile_h = th1 + th2 + int(H * 0.08)
            step_x = tile_w
            step_y = tile_h
            y = -step_y
            while y < H + step_y:
                x = -step_x
                while x < W + step_x:
                    draw.text((x, y), 'Caryanams', fill=color1, font=font1)
                    draw.text((x + (tw1 - tw2)//2, y + th1 + 4), 'Driven by Trust', fill=color2, font=font2)
                    x += step_x
                y += step_y
            blurred = _PILImg.alpha_composite(blurred.convert('RGBA'), wm_layer).convert('RGB')
        except Exception as wm_e:
            current_app.logger.warning(f'[admin blur_bg] watermark failed (non-fatal): {wm_e}')
        blurred.save(out_path, 'JPEG', quality=95)
    except Exception as e:
        current_app.logger.error(f'[admin blur_bg] blur failed: {e}')
        try:
            _PILImg.open(src_path).convert('RGB').save(out_path, 'JPEG', quality=95)
        except Exception:
            return jsonify({'success': False, 'error': 'Processing failed'}), 500

    return jsonify({'success': True, 'processed_url': '/static/processed/' + out_fname, 'method': method})


@admin_bp.route('/api/vehicles/images/<int:image_id>/apply-plate', methods=['POST'])
@require_permission_api('vehicles')
def api_vehicle_apply_plate(image_id):
    """
    Hide number plate on a vehicle image (manual rect or 4-pt quad).
    If use_processed=True, applies on top of most recent temp blur for this image.
    """
    from models import VehicleImage
    from background.utils import detect_number_plate, apply_plate_removal

    vi = VehicleImage.query.get_or_404(image_id)
    data          = request.get_json(silent=True) or {}
    mode          = data.get('mode', 'caryanams')
    manual        = data.get('manual')
    quad          = data.get('quad')
    use_processed = data.get('use_processed', False)

    src_path = os.path.join(current_app.root_path, 'static', 'images', 'uploads', vi.filename)
    if use_processed:
        pf = os.path.join(current_app.root_path, 'static', 'processed')
        candidates = [os.path.join(pf, fn) for fn in os.listdir(pf)
                      if fn.startswith(f'vblur_{image_id}_')] if os.path.isdir(pf) else []
        if candidates:
            src_path = max(candidates, key=os.path.getmtime)

    if not os.path.exists(src_path):
        return jsonify({'success': False, 'message': 'Source file not found'}), 404

    if manual:
        plate = (int(manual.get('x', 0)), int(manual.get('y', 0)),
                 int(manual.get('w', 0)), int(manual.get('h', 0)))
    else:
        plate = detect_number_plate(src_path)

    if not plate:
        return jsonify({'success': False, 'message': 'No plate detected. Use draw mode.'}), 400

    pf = os.path.join(current_app.root_path, 'static', 'processed')
    os.makedirs(pf, exist_ok=True)
    out_fname = f'vplate_{image_id}_{uuid.uuid4().hex[:8]}.png'
    out_path  = os.path.join(pf, out_fname)
    ok = apply_plate_removal(src_path, out_path, *plate, mode=mode, quad=quad)
    if ok and os.path.exists(out_path):
        return jsonify({
            'success':       True,
            'processed_url': '/static/processed/' + out_fname,
            'plate':         {'x': plate[0], 'y': plate[1], 'width': plate[2], 'height': plate[3]},
            'message':       '✅ Plate hidden!'
        })
    return jsonify({'success': False, 'message': 'Plate removal failed.'}), 500


@admin_bp.route('/api/vehicles/images/<int:image_id>/save-processed', methods=['POST'])
@require_permission_api('vehicles')
def api_vehicle_save_processed(image_id):
    """
    Copy a temp processed image (blur/plate) back to the VehicleImage file,
    replacing it in-place so the DB filename stays unchanged.
    Body: { processed_url: '/static/processed/vblur_xxx.jpg' }
    """
    from models import VehicleImage, Vehicle
    from extensions import db as _db
    from PIL import Image as _PILImg
    import shutil

    vi = VehicleImage.query.get_or_404(image_id)
    data          = request.get_json(silent=True) or {}
    processed_url = data.get('processed_url', '')
    if not processed_url:
        return jsonify({'success': False, 'error': 'processed_url required'}), 400

    url_path = processed_url.split('?')[0].lstrip('/')
    src_path = os.path.join(current_app.root_path, url_path)
    if not os.path.exists(src_path):
        return jsonify({'success': False, 'error': 'Processed file not found on disk'}), 404

    dest_path = os.path.join(current_app.root_path, 'static', 'images', 'uploads', vi.filename)
    ext_dest  = os.path.splitext(vi.filename)[1].lower()
    try:
        if ext_dest in ('.jpg', '.jpeg'):
            _PILImg.open(src_path).convert('RGB').save(dest_path, 'JPEG', quality=92)
        else:
            shutil.copy2(src_path, dest_path)
    except Exception as e:
        current_app.logger.error(f'[admin save_processed] failed: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500

    try:
        v = Vehicle.query.get(vi.vehicle_id)
        if v and v.image_filename == vi.filename:
            v.updated_at = _now_ist()
            _db.session.commit()
    except Exception:
        pass

    log_admin_action(f'Saved processed image for vehicle image {image_id}', 'Vehicles')
    return jsonify({
        'success':   True,
        'vi_id':     image_id,
        'image_url': f'/static/images/uploads/{vi.filename}?t={uuid.uuid4().hex[:8]}'
    })


# ══════════════════════════════════════════════════════════════════════════════
# SUB-ADMIN MANAGEMENT
# ══════════════════════════════════════════════════════════════════════════════

ALL_PERMISSIONS = [
    ('dealers',   'Dealers',   'fas fa-handshake'),
    ('vehicles',  'Vehicles',  'fas fa-car'),
    ('leads',     'Leads',     'fas fa-user-tag'),
    ('kyc',       'KYC',       'fas fa-id-card'),
    ('users',     'Users',     'fas fa-users'),
    ('documents', 'Documents', 'fas fa-database'),
    ('reports',   'Reports',   'fas fa-chart-bar'),
    ('settings',  'Settings',  'fas fa-cog'),
]


@admin_bp.route('/sub-admins')
@super_admin_only
def sub_admins():
    from models import SubAdmin
    admins = SubAdmin.query.order_by(SubAdmin.created_at.desc()).all()
    return render_template('admin/sub_admins.html', sub_admins=admins, all_perms=ALL_PERMISSIONS)


@admin_bp.route('/sub-admins/add', methods=['GET', 'POST'])
@super_admin_only
def add_sub_admin():
    from models import SubAdmin
    from extensions import db

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        username = request.form.get('username', '').strip()
        phone = request.form.get('phone', '').strip()
        password = request.form.get('password', '').strip()
        perms = ','.join(request.form.getlist('permissions'))

        if SubAdmin.query.filter_by(username=username).first():
            flash('Username already taken.', 'error')
            return redirect(url_for('admin.add_sub_admin'))
        if SubAdmin.query.filter_by(email=email).first():
            flash('Email already registered.', 'error')
            return redirect(url_for('admin.add_sub_admin'))

        sa = SubAdmin(
            name=name, email=email, username=username, phone=phone,
            permissions=perms, is_active=True,
            created_by=session.get('admin_username', 'admin')
        )
        sa.set_password(password)
        db.session.add(sa)
        db.session.flush()  # get sa.id before commit
        from models import generate_display_id
        sa.display_id = generate_display_id('sub_admin')
        db.session.commit()
        log_admin_action(f'Created sub-admin: {username}', 'SubAdmin')
        flash(f'Sub-admin "{name}" created successfully.', 'success')
        return redirect(url_for('admin.sub_admins'))

    return render_template('admin/add_sub_admin.html', all_perms=ALL_PERMISSIONS)


@admin_bp.route('/sub-admins/<int:sa_id>/edit', methods=['GET', 'POST'])
@super_admin_only
def edit_sub_admin(sa_id):
    from models import SubAdmin
    from extensions import db

    sa = SubAdmin.query.get_or_404(sa_id)

    if request.method == 'POST':
        sa.name = request.form.get('name', sa.name).strip()
        sa.email = request.form.get('email', sa.email).strip()
        sa.phone = request.form.get('phone', '').strip()
        sa.permissions = ','.join(request.form.getlist('permissions'))
        new_pw = request.form.get('password', '').strip()
        if new_pw:
            sa.set_password(new_pw)
        db.session.commit()
        log_admin_action(f'Updated sub-admin: {sa.username}', 'SubAdmin')
        flash('Sub-admin updated.', 'success')
        return redirect(url_for('admin.sub_admins'))

    return render_template('admin/edit_sub_admin.html', sa=sa, all_perms=ALL_PERMISSIONS)


@admin_bp.route('/api/sub-admins/<int:sa_id>/toggle', methods=['POST'])
@super_admin_only_api
def toggle_sub_admin(sa_id):
    from models import SubAdmin
    from extensions import db

    sa = SubAdmin.query.get_or_404(sa_id)
    sa.is_active = not sa.is_active
    db.session.commit()
    log_admin_action(f'{"Activated" if sa.is_active else "Deactivated"} sub-admin: {sa.username}', 'SubAdmin')
    return jsonify({'success': True, 'is_active': sa.is_active})


@admin_bp.route('/api/sub-admins/<int:sa_id>/delete', methods=['POST'])
@super_admin_only_api
def delete_sub_admin(sa_id):
    from models import SubAdmin
    from extensions import db

    sa = SubAdmin.query.get_or_404(sa_id)
    username = sa.username
    db.session.delete(sa)
    db.session.commit()
    log_admin_action(f'Deleted sub-admin: {username}', 'SubAdmin')
    return jsonify({'success': True})


# ─── Sub-admin Login ──────────────────────────────────────────────────────────

# ═══════════════════════════════════════════════════════════════════════════════
# CENTRALIZED DOCUMENT STORAGE — ADMIN MODULE
# ═══════════════════════════════════════════════════════════════════════════════

@admin_bp.route('/document-storage')
@require_permission('documents')
def document_storage():
    """Admin: Centralized Document Storage dashboard."""
    from db import cds_list_all
    from models import User

    filters = {
        'status':      request.args.get('status', 'active'),
        'module_name': request.args.get('module', ''),
        'dealer_id':   request.args.get('dealer_id', '', type=str),
        'search':      request.args.get('q', ''),
    }
    if filters['dealer_id']:
        try:
            filters['dealer_id'] = int(filters['dealer_id'])
        except ValueError:
            filters['dealer_id'] = None

    records  = cds_list_all(filters)
    # All active records (unfiltered) for the dealer summary table
    all_active_records = cds_list_all({'status': 'active'})
    dealers  = User.query.filter_by(role='dealer', is_active=True).order_by(User.name).all()

    from models import CentralDocumentStorage
    total       = CentralDocumentStorage.query.count()
    active_cnt  = CentralDocumentStorage.query.filter_by(status='active').count()
    deleted_cnt = CentralDocumentStorage.query.filter_by(status='deleted').count()
    stats = {'total': total, 'active': active_cnt, 'deleted': deleted_cnt}

    return render_template(
        'admin/document_storage.html',
        records=records,
        all_active_records=all_active_records,
        dealers=dealers,
        filters=filters,
        stats=stats,
        page='document_storage',
    )


@admin_bp.route('/document-storage/upload', methods=['POST'])
@super_admin_only
def cds_upload_doc():
    """Admin: Upload a new document and assign directly to a dealer."""
    import uuid as _uuid
    from db import cds_register
    from models import User

    file      = request.files.get('file')
    dealer_id = request.form.get('dealer_id', type=int)
    module    = request.form.get('module_name', 'Documents').strip()
    doc_type  = request.form.get('document_type', '').strip()
    notes     = request.form.get('notes', '').strip()

    if not file or not file.filename:
        flash('Please select a file to upload.', 'error')
        return redirect(url_for('admin.document_storage'))
    if not dealer_id:
        flash('Please select a dealer to assign this document to.', 'error')
        return redirect(url_for('admin.document_storage'))

    dealer = User.query.get(dealer_id)
    if not dealer or dealer.role != 'dealer':
        flash('Invalid dealer selected.', 'error')
        return redirect(url_for('admin.document_storage'))

    ext      = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else 'bin'
    filename = f"{_uuid.uuid4().hex}.{ext}"
    rel_path = os.path.join('images', 'uploads', filename)
    full_path = os.path.join(current_app.root_path, 'static', rel_path)
    os.makedirs(os.path.dirname(full_path), exist_ok=True)
    file.save(full_path)

    actor, role_label = _resolve_role_and_user()
    cds_register({
        'dealer_id':     dealer_id,
        'file_name':     filename,
        'original_name': file.filename,
        'file_path':     rel_path,
        'module_name':   module,
        'document_type': doc_type,
        'uploaded_by':   None,
        'performed_by':  actor,
        'user_role':     role_label,
    })
    log_admin_action(f"Uploaded doc '{file.filename}' to dealer {dealer.name}", 'Documents')
    flash(f'Document "{file.filename}" uploaded and assigned to {dealer.name} successfully.', 'success')
    return redirect(url_for('admin.document_storage'))


@admin_bp.route('/document-storage/<int:record_id>/download')
@require_permission('documents')
def cds_download(record_id):
    """Admin: Download any file from centralized storage."""
    import mimetypes
    from flask import send_file, abort
    from db import cds_get

    rec = cds_get(record_id)
    if not rec:
        abort(404)

    full_path = os.path.join(current_app.root_path, 'static', rec.file_path)
    if not os.path.exists(full_path):
        full_path = os.path.join(current_app.config.get('UPLOAD_FOLDER', ''), rec.file_name)
        if not os.path.exists(full_path):
            flash('File not found on disk.', 'error')
            return redirect(url_for('admin.document_storage'))

    mime, _ = mimetypes.guess_type(full_path)
    return send_file(
        full_path,
        as_attachment=True,
        download_name=rec.original_name or rec.file_name,
        mimetype=mime or 'application/octet-stream',
    )


@admin_bp.route('/document-storage/<int:record_id>/soft-delete', methods=['POST'])
@super_admin_only
def cds_soft_delete_doc(record_id):
    """Admin: Soft-delete a document (marks deleted, file stays on disk)."""
    from db import cds_soft_delete
    actor, role_label = _resolve_role_and_user()
    ok = cds_soft_delete(record_id, performed_by=actor, user_role=role_label)
    if ok:
        flash('Document marked as deleted. It is no longer visible to the dealer.', 'success')
    else:
        flash('Delete failed — document not found or already deleted.', 'error')
    return redirect(url_for('admin.document_storage'))


@admin_bp.route('/document-storage/<int:record_id>/delete', methods=['POST'])
@super_admin_only
def cds_delete_doc(record_id):
    """Admin: Permanently delete file from storage."""
    from db import cds_hard_delete
    actor = session.get('admin_username') or session.get('sub_admin_username', 'admin')
    upload_folder = current_app.config.get('UPLOAD_FOLDER', '')
    ok = cds_hard_delete(record_id, upload_folder=upload_folder, performed_by=actor)
    if ok:
        flash('File permanently deleted.', 'success')
    else:
        flash('File not found.', 'error')
    return redirect(url_for('admin.document_storage'))


@admin_bp.route('/document-storage/<int:record_id>/audit')
@require_permission('documents')
def cds_audit(record_id):
    """Admin: View audit log for a specific document."""
    from db import cds_get, cds_get_audit_logs
    rec  = cds_get(record_id)
    logs = cds_get_audit_logs(record_id)
    return render_template(
        'admin/document_storage_audit.html',
        record=rec,
        logs=logs,
        page='document_storage',
    )


# ─── Sub-admin Login ──────────────────────────────────────────────────────────
@admin_bp.route('/sub-admin-login', methods=['GET', 'POST'])
def sub_admin_login():
    if session.get('admin_logged_in') or session.get('sub_admin_logged_in'):
        return redirect(url_for('admin.dashboard'))

    if request.method == 'POST':
        from models import SubAdmin
        from datetime import datetime
        from extensions import db

        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        sa = SubAdmin.query.filter_by(username=username, is_active=True).first()
        if sa and sa.check_password(password):
            session['sub_admin_logged_in'] = True
            session['sub_admin_id'] = sa.id
            session['sub_admin_username'] = sa.username
            session['sub_admin_name'] = sa.name
            session['sub_admin_permissions'] = sa.get_permissions()
            sa.last_login = _now_ist()
            db.session.commit()
            log_admin_action('Sub Admin logged in', 'Auth')
            flash(f'Welcome, {sa.name}!', 'success')
            return redirect(url_for('admin.dashboard'))
        # Log failed sub-admin login attempt
        try:
            from models import AdminLog
            from utils.request_meta import get_request_meta
            ip, browser, os_name, device = get_request_meta(request)
            db.session.add(AdminLog(
                user_id=sa.id if sa else None,
                admin_user=username or 'unknown',
                user_role='Sub Admin',
                action='Failed sub-admin login attempt',
                module='Auth',
                description=f'Failed login attempt for username "{username}"',
                ip_address=ip,
                device=device,
                browser=browser,
                timezone='Asia/Kolkata (IST)',
                status='Failed',
            ))
            db.session.commit()
        except Exception:
            pass
        flash('Invalid credentials or account inactive.', 'error')

    return render_template('admin/sub_admin_login.html')


# ═══════════════════════════════════════════════════════════════════════════════
# LEAD IMPORT & ASSIGNMENT MODULE
# ═══════════════════════════════════════════════════════════════════════════════

import json as _json

LEAD_UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), '..', 'static', 'uploads', 'lead_imports')
os.makedirs(LEAD_UPLOAD_FOLDER, exist_ok=True)
ALLOWED_LEAD_EXTENSIONS = {'csv', 'xlsx', 'xls'}


def _allowed_lead_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_LEAD_EXTENSIONS


def _parse_lead_file(filepath, ext):
    """
    Parse CSV or Excel file. Returns list of row dicts.
    Raises ValueError on unreadable files.
    """
    rows = []
    if ext == 'csv':
        import csv as _csv
        with open(filepath, newline='', encoding='utf-8-sig', errors='replace') as f:
            reader = _csv.DictReader(f)
            for row in reader:
                rows.append(dict(row))
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c).strip() if c is not None else f'col_{j}' for j, c in enumerate(row)]
                else:
                    if all(c is None for c in row):
                        continue
                    rows.append(dict(zip(headers, [str(c).strip() if c is not None else '' for c in row])))
            wb.close()
        except ImportError:
            raise ValueError("openpyxl not installed. Run: pip install openpyxl")
    return rows


def _normalize_row(row):
    """
    Map flexible column names to standard fields.
    Returns dict with: name, phone, email, company, address, source, extra
    """
    def _get(row, *keys):
        for k in keys:
            for rk in row:
                if rk.strip().lower().replace(' ', '_') == k.lower().replace(' ', '_'):
                    v = str(row[rk]).strip()
                    if v and v.lower() not in ('none', 'nan', 'null', ''):
                        return v
        return None

    KNOWN_KEYS = {'name', 'customer_name', 'full_name',
                  'phone', 'phone_number', 'mobile', 'contact',
                  'email', 'email_address',
                  'company', 'company_name', 'organization',
                  'address', 'location',
                  'source', 'lead_source'}
    name    = _get(row, 'name', 'customer_name', 'full_name', 'Name', 'Customer Name')
    phone   = _get(row, 'phone', 'phone_number', 'mobile', 'contact', 'Mobile', 'Phone')
    email   = _get(row, 'email', 'email_address', 'Email')
    company = _get(row, 'company', 'company_name', 'organization', 'Company')
    address = _get(row, 'address', 'location', 'Address')
    source  = _get(row, 'source', 'lead_source', 'Source') or 'Import'

    # capture unknown columns as extra
    extra = {}
    for k, v in row.items():
        norm = k.strip().lower().replace(' ', '_')
        if norm not in KNOWN_KEYS and str(v).strip() not in ('', 'none', 'nan', 'null'):
            extra[k.strip()] = str(v).strip()

    return name, phone, email, company, address, source, extra


# ── Page: Import Leads ────────────────────────────────────────────────────────

@admin_bp.route('/leads/import-file', methods=['GET', 'POST'])
@require_permission('leads')
def lead_import_page():
    from models import LeadImportFile, ImportedLead
    from extensions import db

    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected.', 'error')
            return redirect(url_for('admin.lead_import_page'))

        f = request.files['file']
        if not f.filename:
            flash('No file selected.', 'error')
            return redirect(url_for('admin.lead_import_page'))

        if not _allowed_lead_file(f.filename):
            flash('Unsupported file format. Please upload CSV, XLSX, or XLS.', 'error')
            return redirect(url_for('admin.lead_import_page'))

        from werkzeug.utils import secure_filename
        ext = f.filename.rsplit('.', 1)[1].lower()
        stored_name = f'{uuid.uuid4().hex}.{ext}'
        filepath = os.path.join(LEAD_UPLOAD_FOLDER, stored_name)
        f.save(filepath)

        # create import record
        import_record = LeadImportFile(
            file_name=secure_filename(f.filename),
            stored_name=stored_name,
            file_type=ext,
            uploaded_by=session.get('admin_username') or session.get('sub_admin_username') or 'admin',
            status='processing'
        )
        db.session.add(import_record)
        db.session.commit()

        # parse file
        try:
            rows = _parse_lead_file(filepath, ext)
        except Exception as e:
            import_record.status = 'failed'
            import_record.error_message = str(e)
            db.session.commit()
            flash(f'File parse error: {e}', 'error')
            return redirect(url_for('admin.lead_import_page'))

        import_record.total_rows = len(rows)
        imported = 0
        duplicates = 0
        failed = 0
        leads_to_insert = []

        for row in rows:
            try:
                name, phone, email, company, address, source, extra = _normalize_row(row)

                if not name or not phone:
                    failed += 1
                    continue

                # duplicate check: same phone or same email
                dup = ImportedLead.query.filter(
                    (ImportedLead.phone == phone) |
                    ((ImportedLead.email == email) if email else False)
                ).first()
                if dup:
                    duplicates += 1
                    continue

                lead = ImportedLead(
                    import_file_id=import_record.id,
                    name=name, phone=phone, email=email,
                    company=company, address=address, source=source,
                    extra_data=_json.dumps(extra) if extra else None,
                    status='New'
                )
                leads_to_insert.append(lead)
            except Exception:
                failed += 1

        # bulk insert
        if leads_to_insert:
            db.session.bulk_save_objects(leads_to_insert)
            imported = len(leads_to_insert)

        import_record.imported_rows = imported
        import_record.duplicate_rows = duplicates
        import_record.failed_rows = failed
        import_record.status = 'done'
        db.session.commit()

        log_admin_action(f"Imported {imported} leads from '{import_record.file_name}'", 'Leads')
        flash(f'✅ Import complete! {imported} leads imported, {duplicates} duplicates skipped, {failed} failed.', 'success')
        return redirect(url_for('admin.imported_leads_list'))

    # GET — show import page with history
    import_history = LeadImportFile.query.order_by(LeadImportFile.uploaded_at.desc()).limit(20).all()
    return render_template('admin/import_leads.html',
                           import_history=import_history, page='leads')


# ── Page: Imported Leads List ─────────────────────────────────────────────────

@admin_bp.route('/leads/imported')
@require_permission('leads')
def imported_leads_list():
    from models import ImportedLead, LeadImportFile, User as _User
    from extensions import db

    # filters
    search      = request.args.get('search', '').strip()
    status_f    = request.args.get('status', '').strip()
    dealer_f    = request.args.get('dealer_id', '').strip()
    import_f    = request.args.get('import_id', '').strip()
    date_from   = request.args.get('date_from', '').strip()
    date_to     = request.args.get('date_to', '').strip()
    sort_col    = request.args.get('sort', 'created_at')
    sort_dir    = request.args.get('dir', 'desc')
    page        = request.args.get('page', 1, type=int)
    per_page    = 25
    export      = request.args.get('export', '')

    query = ImportedLead.query

    if search:
        like = f'%{search}%'
        query = query.filter(
            (ImportedLead.name.ilike(like)) |
            (ImportedLead.phone.ilike(like)) |
            (ImportedLead.email.ilike(like)) |
            (ImportedLead.company.ilike(like))
        )
    if status_f:
        query = query.filter(ImportedLead.status == status_f)
    if dealer_f:
        if dealer_f == 'unassigned':
            query = query.filter(ImportedLead.assigned_dealer_id.is_(None))
        else:
            query = query.filter(ImportedLead.assigned_dealer_id == int(dealer_f))
    if import_f:
        query = query.filter(ImportedLead.import_file_id == int(import_f))
    if date_from:
        try:
            query = query.filter(ImportedLead.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
        except Exception:
            pass
    if date_to:
        try:
            from datetime import timedelta
            query = query.filter(ImportedLead.created_at < datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))
        except Exception:
            pass

    # sort
    sort_map = {
        'name':       ImportedLead.name,
        'phone':      ImportedLead.phone,
        'status':     ImportedLead.status,
        'created_at': ImportedLead.created_at,
        'assigned_at':ImportedLead.assigned_at,
    }
    sort_attr = sort_map.get(sort_col, ImportedLead.created_at)
    query = query.order_by(sort_attr.desc() if sort_dir == 'desc' else sort_attr.asc())

    total = query.count()

    # CSV export
    if export == '1':
        all_leads = query.all()
        import csv as _csv
        import io as _io
        buf = _io.StringIO()
        w = _csv.writer(buf)
        w.writerow(['ID', 'Name', 'Phone', 'Email', 'Company', 'Address',
                    'Source', 'Status', 'Assigned Dealer', 'Assigned At',
                    'Import File', 'Created At'])
        for l in all_leads:
            w.writerow([
                l.id, l.name, l.phone, l.email or '', l.company or '',
                l.address or '', l.source, l.status,
                l.assigned_dealer.name if l.assigned_dealer else '',
                l.assigned_at.strftime('%d %b %Y %H:%M') if l.assigned_at else '',
                l.import_file.file_name if l.import_file else '',
                l.created_at.strftime('%d %b %Y %H:%M') if l.created_at else '',
            ])
        output = make_response(buf.getvalue())
        output.headers['Content-Type'] = 'text/csv'
        output.headers['Content-Disposition'] = 'attachment; filename=imported_leads.csv'
        return output

    leads        = query.offset((page - 1) * per_page).limit(per_page).all()
    total_pages  = (total + per_page - 1) // per_page
    dealers      = _User.query.filter_by(role='dealer', is_active=True).order_by(_User.name).all()
    import_files = LeadImportFile.query.order_by(LeadImportFile.uploaded_at.desc()).all()

    # dashboard stats
    stats = {
        'total':      ImportedLead.query.count(),
        'assigned':   ImportedLead.query.filter(ImportedLead.assigned_dealer_id.isnot(None)).count(),
        'unassigned': ImportedLead.query.filter(ImportedLead.assigned_dealer_id.is_(None)).count(),
        'today':      ImportedLead.query.filter(
            ImportedLead.created_at >= _now_ist().replace(hour=0, minute=0, second=0, microsecond=0)
        ).count(),
    }

    return render_template('admin/imported_leads.html',
                           leads=leads, dealers=dealers, import_files=import_files,
                           total=total, page=page, total_pages=total_pages,
                           per_page=per_page, stats=stats,
                           search=search, status_f=status_f, dealer_f=dealer_f,
                           import_f=import_f, date_from=date_from, date_to=date_to,
                           sort_col=sort_col, sort_dir=sort_dir,
                           page_name='leads')


# ── API: Single Lead Assignment ───────────────────────────────────────────────

@admin_bp.route('/leads/imported/<int:lead_id>/assign', methods=['POST'])
@admin_api_login_required
def assign_imported_lead(lead_id):
    from models import ImportedLead, LeadAssignmentHistory, User as _User
    from extensions import db

    lead = ImportedLead.query.get_or_404(lead_id)
    data = request.get_json(silent=True) or {}
    dealer_id = data.get('dealer_id')

    if not dealer_id:
        return jsonify({'success': False, 'message': 'dealer_id required'}), 400

    dealer = _User.query.filter_by(id=dealer_id, role='dealer').first()
    if not dealer:
        return jsonify({'success': False, 'message': 'Dealer not found'}), 404

    action = 'reassigned' if lead.assigned_dealer_id else 'assigned'
    lead.assigned_dealer_id = dealer.id
    lead.assigned_at  = _now_ist()
    lead.assigned_by  = session.get('admin_username') or session.get('sub_admin_username') or 'admin'
    lead.status       = 'Assigned'

    hist = LeadAssignmentHistory(
        lead_id=lead.id, dealer_id=dealer.id,
        dealer_name=dealer.name, action=action,
        assigned_by=lead.assigned_by
    )
    db.session.add(hist)
    db.session.commit()

    log_admin_action(f"{action.capitalize()} lead #{lead.id} ({lead.name}) to {dealer.name}", 'Leads')
    return jsonify({'success': True, 'message': f'Lead assigned to {dealer.name}',
                    'dealer_name': dealer.name, 'assigned_at': lead.assigned_at.strftime('%d %b %Y %H:%M')})


# ── API: Bulk Lead Assignment ─────────────────────────────────────────────────

@admin_bp.route('/leads/imported/bulk-assign', methods=['POST'])
@admin_api_login_required
def bulk_assign_imported_leads():
    from models import ImportedLead, LeadAssignmentHistory, User as _User
    from extensions import db

    data      = request.get_json(silent=True) or {}
    lead_ids  = data.get('lead_ids', [])
    dealer_id = data.get('dealer_id')

    if not lead_ids or not dealer_id:
        return jsonify({'success': False, 'message': 'lead_ids and dealer_id required'}), 400

    dealer = _User.query.filter_by(id=dealer_id, role='dealer').first()
    if not dealer:
        return jsonify({'success': False, 'message': 'Dealer not found'}), 404

    admin_user = session.get('admin_username') or session.get('sub_admin_username') or 'admin'
    now = _now_ist()
    updated = 0
    hist_records = []

    leads = ImportedLead.query.filter(ImportedLead.id.in_(lead_ids)).all()
    for lead in leads:
        action = 'reassigned' if lead.assigned_dealer_id else 'assigned'
        lead.assigned_dealer_id = dealer.id
        lead.assigned_at  = now
        lead.assigned_by  = admin_user
        lead.status       = 'Assigned'
        hist_records.append(LeadAssignmentHistory(
            lead_id=lead.id, dealer_id=dealer.id,
            dealer_name=dealer.name, action=action, assigned_by=admin_user
        ))
        updated += 1

    db.session.add_all(hist_records)
    db.session.commit()

    log_admin_action(f"Bulk assigned {updated} leads to {dealer.name}", 'Leads')
    return jsonify({'success': True, 'message': f'{updated} leads assigned to {dealer.name}', 'count': updated})


# ── API: Update Lead Status ───────────────────────────────────────────────────

@admin_bp.route('/leads/imported/<int:lead_id>/status', methods=['POST'])
@admin_api_login_required
def update_imported_lead_status(lead_id):
    from models import ImportedLead
    from extensions import db

    lead = ImportedLead.query.get_or_404(lead_id)
    data = request.get_json(silent=True) or {}
    new_status = data.get('status', '').strip()

    VALID = ['New', 'Assigned', 'Contacted', 'Follow-up', 'Converted', 'Rejected']
    if new_status not in VALID:
        return jsonify({'success': False, 'message': f'Invalid status. Must be one of: {VALID}'}), 400

    lead.status = new_status
    db.session.commit()
    log_admin_action(f"Updated imported lead #{lead.id} status to {new_status}", 'Leads')
    return jsonify({'success': True, 'message': f'Status updated to {new_status}'})


# ── API: Lead Assignment History ──────────────────────────────────────────────

@admin_bp.route('/leads/imported/<int:lead_id>/history')
@admin_api_login_required
def imported_lead_history(lead_id):
    from models import ImportedLead, LeadAssignmentHistory
    lead = ImportedLead.query.get_or_404(lead_id)
    history = LeadAssignmentHistory.query.filter_by(lead_id=lead.id)\
                  .order_by(LeadAssignmentHistory.assigned_at.desc()).all()
    return jsonify({'success': True, 'lead': lead.to_dict(), 'history': [
        {'id': h.id, 'dealer_name': h.dealer_name, 'action': h.action,
         'assigned_by': h.assigned_by,
         'assigned_at': h.assigned_at.strftime('%d %b %Y %H:%M') if h.assigned_at else ''}
        for h in history
    ]})


# ── API: Delete Imported Lead ─────────────────────────────────────────────────

@admin_bp.route('/leads/imported/<int:lead_id>/delete', methods=['POST'])
@admin_api_login_required
def delete_imported_lead(lead_id):
    from models import ImportedLead
    from extensions import db
    lead = ImportedLead.query.get_or_404(lead_id)
    name = lead.name
    db.session.delete(lead)
    db.session.commit()
    log_admin_action(f"Deleted imported lead #{lead_id} ({name})", 'Leads')
    return jsonify({'success': True, 'message': 'Lead deleted'})


# ── API: Dashboard Stats ──────────────────────────────────────────────────────

@admin_bp.route('/api/leads/import-stats')
@admin_api_login_required
def lead_import_stats():
    from models import ImportedLead, User as _User
    from extensions import db
    from sqlalchemy import func

    total     = ImportedLead.query.count()
    assigned  = ImportedLead.query.filter(ImportedLead.assigned_dealer_id.isnot(None)).count()
    today_start = _now_ist().replace(hour=0, minute=0, second=0, microsecond=0)
    today     = ImportedLead.query.filter(ImportedLead.created_at >= today_start).count()

    dealer_counts = db.session.query(
        _User.name, func.count(ImportedLead.id).label('cnt')
    ).join(ImportedLead, ImportedLead.assigned_dealer_id == _User.id)\
     .group_by(_User.id).order_by(func.count(ImportedLead.id).desc()).limit(10).all()

    return jsonify({
        'total': total, 'assigned': assigned, 'unassigned': total - assigned,
        'today': today,
        'dealer_counts': [{'name': r[0], 'count': r[1]} for r in dealer_counts]
    })
